from fastapi import FastAPI, Depends, HTTPException, status
from fastapi.security import OAuth2PasswordRequestForm
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import io
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pydantic import BaseModel
from datetime import date, timedelta, datetime
import secrets
import os
import uuid
import shutil
from fastapi import UploadFile, File, Form
from fastapi.staticfiles import StaticFiles
from dotenv import load_dotenv
# Charger les variables d'environnement
load_dotenv()
RESEND_API_KEY = os.getenv("RESEND_API_KEY")

from database import engine, SessionLocal, Base
from models import (
    Stagiaire,
    Etablissement,
    Departement,
    Encadrant,
    Utilisateur,
    Document,
    Presence,
    ConfirmationEvenement,
    Activite,
    Formation,
    Competence,
    SessionFormation,
    CommentaireEncadrant,
    Evaluation,
    Message,
    Reunion,
    Notification,
    DemandeStage,
    SujetPFE,
)
from auth import (
    hasher_mot_de_passe,
    verifier_mot_de_passe,
    creer_token_acces,
    obtenir_utilisateur_courant,
    exiger_role,
    get_db,
)
from schemas import ProfilUpdate, ChangementMotDePasse, MotDePasseOublie, ReinitialiserMotDePasse, CompetenceCreate, CompetenceOut, FormationCreate, FormationOut, CommentaireCreate, CommentaireOut, DatesUpdate, IncidentCreate, StagiaireDisponibleOut, DemandeStageCreate, DemandeStageUpdate, SujetPFECreate, SujetPFEUpdate, SujetPFEResponse

# Crée toutes les tables si elles n'existent pas encore
Base.metadata.create_all(bind=engine)

app = FastAPI(title="API Suivi des Stagiaires")

UPLOAD_DIR = "uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)
app.mount("/uploads", StaticFiles(directory=UPLOAD_DIR), name="uploads")

# Stockage en mémoire des tokens de réinitialisation (simulation, non persistant)
tokens_reinitialisation = {}

from fastapi.middleware.cors import CORSMiddleware

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://localhost:5500"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ============================================================
# Connexion à la base de données (dépendance réutilisée partout)
# ============================================================
# get_db() est importée depuis auth.py pour garantir une seule
# session par requête (évite les bugs de session "not persistent")


# ============================================================
# Schémas Pydantic (validation des données reçues)
# ============================================================

class EtablissementCreate(BaseModel):
    nom: str


class DepartementCreate(BaseModel):
    nom: str


class EncadrantCreate(BaseModel):
    nom: str
    prenom: str
    email: str | None = None
    departement_id: int | None = None


class UtilisateurCreate(BaseModel):
    nom: str
    email: str
    mot_de_passe_hash: str
    stagiaire_id: int | None = None
    encadrant_id: int | None = None
    role: str


class CreationCompteStagiaire(BaseModel):
    stagiaire_id: int
    email: str
    mot_de_passe: str


class StagiaireCreate(BaseModel):
    prenom: str
    nom: str
    email: str | None = None
    telephone: str | None = None
    cin: str | None = None
    etablissement_id: int | None = None
    niveau_etudes: str | None = None
    specialisation: str | None = None
    type_stage: str | None = None
    departement_id: int | None = None
    encadrant_id: int | None = None
    date_debut: date
    date_fin: date


class DocumentCreate(BaseModel):
    stagiaire_id: int
    type_document: str
    statut: str | None = "en_attente"
    date_document: date | None = None
    fichier_url: str | None = None


def formater_taille(taille_octets):
    if not taille_octets:
        return None
    if taille_octets < 1024 * 1024:
        return f"{taille_octets / 1024:.0f} KB"
    return f"{taille_octets / (1024 * 1024):.1f} MB"


class DocumentOut(BaseModel):
    id: int
    nom: str
    statut: str
    icon: str
    valide: bool
    commentaire: str | None = None
    date_document: date | None = None
    fichier_url: str | None = None
    taille_affichee: str | None = None
    origine: str = "stagiaire"

    class Config:
        from_attributes = True

    @classmethod
    def depuis_document(cls, doc):
        noms = {
            "convention": "Convention de stage",
            "attestation": "Attestation",
            "rapport_intermediaire": "Rapport intermediaire",
            "rapport_final": "Rapport final",
            "lettre_affectation": "Lettre d'affectation",
            "badge_photo": "Badge photo",
            "fiche_securite": "Fiche de securite",
            "certificat": "Certificat",
        }
        icones = {
            "badge_photo": "img",
            "fiche_securite": "doc",
        }
        statuts_valides = {"valide", "genere", "recu"}
        labels_statut = {
            "valide": "Valide",
            "genere": "Valide",
            "recu": "Valide",
            "en_attente": "En attente",
            "manquant": "Manquant",
            "refuse": "Refuse",
            "a_completer": "A completer",
        }
        return cls(
            id=doc.id,
            nom=noms.get(doc.type_document, doc.type_document),
            statut=labels_statut.get(doc.statut, doc.statut),
            icon=icones.get(doc.type_document, "pdf"),
            valide=doc.statut in statuts_valides,
            commentaire=doc.commentaire,
            date_document=doc.date_document,
            fichier_url=doc.fichier_url,
            taille_affichee=formater_taille(doc.taille_octets),
            origine=getattr(doc, "origine", None) or "stagiaire",
        )


class PresenceCreate(BaseModel):
    stagiaire_id: int
    date: date
    present: bool
    marque_par: int | None = None
    heure_arrivee: str | None = None
    heure_depart: str | None = None


class PresenceStagiaireUpdate(BaseModel):
    heure_arrivee: str | None = None
    heure_depart: str | None = None


class ActiviteCreate(BaseModel):
    stagiaire_id: int
    action: str
    statut: str | None = None


class EmailIdentifiantsRequest(BaseModel):
    email: str
    nom: str
    mot_de_passe: str


# ============================================================
# Route d'accueil
# ============================================================

@app.get("/")
def accueil():
    return {"message": "Bienvenue sur l'API de suivi des stagiaires"}


# ============================================================
# Établissements
# ============================================================

@app.get("/etablissements")
def liste_etablissements(db: Session = Depends(get_db)):
    return db.query(Etablissement).all()


@app.post("/etablissements")
def ajouter_etablissement(
    etab: EtablissementCreate,
    db: Session = Depends(get_db),
    utilisateur_courant: Utilisateur = Depends(exiger_role("admin_rh")),
):
    nouveau = Etablissement(**etab.dict())
    db.add(nouveau)
    db.commit()
    db.refresh(nouveau)
    return nouveau


@app.delete("/etablissements/{etablissement_id}")
def supprimer_etablissement(
    etablissement_id: int,
    db: Session = Depends(get_db),
    utilisateur_courant: Utilisateur = Depends(exiger_role("admin_rh")),
):
    obj = db.query(Etablissement).filter(Etablissement.id == etablissement_id).first()
    if obj is None:
        return {"erreur": "Établissement non trouvé"}
    db.delete(obj)
    db.commit()
    return {"message": f"Établissement {etablissement_id} supprimé"}


# ============================================================
# Départements
# ============================================================

@app.get("/departements/export")
def exporter_departements_excel(
    db: Session = Depends(get_db),
    utilisateur_courant: Utilisateur = Depends(exiger_role("rh", "admin_rh")),
):
    departements = db.query(Departement).all()
    stagiaires = db.query(Stagiaire).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Departements"

    entetes = ["Departement", "Total stagiaires", "Stagiaires actifs", "Stagiaires termines"]
    ws.append(entetes)

    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for dept in departements:
        stagiaires_dept = [s for s in stagiaires if s.departement_id == dept.id]
        total = len(stagiaires_dept)
        actifs = len([s for s in stagiaires_dept if s.statut == "en_cours"])
        termines = len([s for s in stagiaires_dept if s.statut == "termine"])
        ws.append([dept.nom, total, actifs, termines])

    largeurs = [30, 18, 18, 20]
    for i, largeur in enumerate(largeurs, start=1):
        ws.column_dimensions[chr(64 + i)].width = largeur

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nom_fichier = f"departements_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nom_fichier}"'},
    )


@app.get("/departements")
def liste_departements(db: Session = Depends(get_db)):
    return db.query(Departement).all()


@app.post("/departements")
def ajouter_departement(
    dep: DepartementCreate,
    db: Session = Depends(get_db),
    utilisateur_courant: Utilisateur = Depends(exiger_role("admin_rh")),
):
    nouveau = Departement(**dep.dict())
    db.add(nouveau)
    db.commit()
    db.refresh(nouveau)
    return nouveau


@app.delete("/departements/{departement_id}")
def supprimer_departement(
    departement_id: int,
    db: Session = Depends(get_db),
    utilisateur_courant: Utilisateur = Depends(exiger_role("admin_rh")),
):
    obj = db.query(Departement).filter(Departement.id == departement_id).first()
    if obj is None:
        return {"erreur": "Département non trouvé"}
    db.delete(obj)
    db.commit()
    return {"message": f"Département {departement_id} supprimé"}


# ============================================================
# Encadrants
# ============================================================

@app.get("/encadrants")
def liste_encadrants(db: Session = Depends(get_db)):
    return db.query(Encadrant).all()


@app.post("/encadrants")
def ajouter_encadrant(
    encadrant: EncadrantCreate,
    db: Session = Depends(get_db),
    utilisateur_courant: Utilisateur = Depends(exiger_role("admin_rh")),
):
    nouveau = Encadrant(**encadrant.dict())
    db.add(nouveau)
    db.commit()
    db.refresh(nouveau)
    return nouveau


@app.delete("/encadrants/{encadrant_id}")
def supprimer_encadrant(
    encadrant_id: int,
    db: Session = Depends(get_db),
    utilisateur_courant: Utilisateur = Depends(exiger_role("admin_rh")),
):
    obj = db.query(Encadrant).filter(Encadrant.id == encadrant_id).first()
    if obj is None:
        return {"erreur": "Encadrant non trouvé"}
    db.delete(obj)
    db.commit()
    return {"message": f"Encadrant {encadrant_id} supprimé"}


# ============================================================
# Utilisateurs
# ============================================================

@app.get("/utilisateurs")
def liste_utilisateurs(
    db: Session = Depends(get_db),
    utilisateur_courant: Utilisateur = Depends(exiger_role("admin_rh")),
):
    return db.query(Utilisateur).all()


@app.post("/utilisateurs")
def ajouter_utilisateur(
    utilisateur: UtilisateurCreate,
    db: Session = Depends(get_db),
    utilisateur_courant: Utilisateur = Depends(exiger_role("admin_rh")),
):
    donnees = utilisateur.dict()
    donnees["mot_de_passe_hash"] = hasher_mot_de_passe(donnees["mot_de_passe_hash"])
    nouveau = Utilisateur(**donnees)
    db.add(nouveau)
    db.commit()
    db.refresh(nouveau)
    return nouveau


@app.post("/stagiaires/{stagiaire_id}/creer-compte")
def creer_compte_stagiaire(
    stagiaire_id: int,
    donnees: CreationCompteStagiaire,
    db: Session = Depends(get_db),
    utilisateur_courant: Utilisateur = Depends(exiger_role("admin_rh")),
):
    # Vérifier que le stagiaire existe
    stagiaire = db.query(Stagiaire).filter(Stagiaire.id == stagiaire_id).first()
    if stagiaire is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Stagiaire non trouvé")

    # Vérifier que ce stagiaire n'a pas déjà un compte
    compte_existant = db.query(Utilisateur).filter(Utilisateur.stagiaire_id == stagiaire_id).first()
    if compte_existant is not None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Ce stagiaire a déjà un compte")

    # Vérifier que l'email n'est pas déjà utilisé
    email_existant = db.query(Utilisateur).filter(Utilisateur.email == donnees.email).first()
    if email_existant is not None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Cet email est déjà utilisé")

    nouveau_compte = Utilisateur(
        nom=f"{stagiaire.prenom} {stagiaire.nom}",
        email=donnees.email,
        mot_de_passe_hash=hasher_mot_de_passe(donnees.mot_de_passe),
        role="stagiaire",
        stagiaire_id=stagiaire_id,
    )
    db.add(nouveau_compte)
    db.commit()
    db.refresh(nouveau_compte)

    return {
        "message": "Compte stagiaire créé avec succès",
        "utilisateur": {
            "id": nouveau_compte.id,
            "nom": nouveau_compte.nom,
            "email": nouveau_compte.email,
            "role": nouveau_compte.role,
            "stagiaire_id": nouveau_compte.stagiaire_id,
        },
    }


@app.delete("/utilisateurs/{utilisateur_id}")
def supprimer_utilisateur(
    utilisateur_id: int,
    db: Session = Depends(get_db),
    utilisateur_courant: Utilisateur = Depends(exiger_role("admin_rh")),
):	
    obj = db.query(Utilisateur).filter(Utilisateur.id == utilisateur_id).first()
    if obj is None:
        return {"erreur": "Utilisateur non trouvé"}
    db.delete(obj)
    db.commit()
    return {"message": f"Utilisateur {utilisateur_id} supprimé"}


@app.get("/stagiaires")
def liste_stagiaires(db: Session = Depends(get_db)):
    return db.query(Stagiaire).all()
@app.get("/stagiaires/export-excel")
def export_stagiaires_excel(db: Session = Depends(get_db)):
    """Exporte la liste des stagiaires en fichier Excel"""
    from openpyxl import Workbook

    stagiaires = db.query(Stagiaire).all()

    etablissements = {e.id: e.nom for e in db.query(Etablissement).all()}
    departements = {d.id: d.nom for d in db.query(Departement).all()}

    wb = Workbook()
    ws = wb.active
    ws.title = "Stagiaires"

    entetes = ["Nom", "Prénom", "Email", "Téléphone", "Établissement", "Département", "Statut", "Date Début", "Date Fin"]
    ws.append(entetes)

    for s in stagiaires:
        ws.append([
            s.nom,
            s.prenom,
            s.email,
            s.telephone,
            etablissements.get(s.etablissement_id, ""),
            departements.get(s.departement_id, ""),
            s.statut,
            s.date_debut.isoformat() if s.date_debut else "",
            s.date_fin.isoformat() if s.date_fin else "",
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=stagiaires_hutchinson.xlsx"}
    )


@app.get("/stagiaires/{stagiaire_id}")
def obtenir_stagiaire(stagiaire_id: int, db: Session = Depends(get_db)):
    stagiaire = db.query(Stagiaire).filter(Stagiaire.id == stagiaire_id).first()
    if stagiaire is None:
        return {"erreur": "Stagiaire non trouvé"}
    return stagiaire


# ============================================================
# Authentification (Login / JWT)
# ============================================================

@app.post("/login")
def login(
    form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)
):
    utilisateur = db.query(Utilisateur).filter(Utilisateur.email == form_data.username).first()
    if utilisateur is None or not verifier_mot_de_passe(form_data.password, utilisateur.mot_de_passe_hash):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Email ou mot de passe incorrect",
        )
    token = creer_token_acces(data={"sub": utilisateur.email})
    return {"access_token": token, "token_type": "bearer"}


@app.get("/moi")
def obtenir_mon_profil(
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    profil = {
        "id": utilisateur_courant.id,
        "nom": utilisateur_courant.nom,
        "email": utilisateur_courant.email,
        "role": utilisateur_courant.role,
        "stage": None,
    }

    if utilisateur_courant.role == "stagiaire" and utilisateur_courant.stagiaire_id:
        stagiaire = db.query(Stagiaire).filter(
            Stagiaire.id == utilisateur_courant.stagiaire_id
        ).first()

        if stagiaire is not None:
            encadrant = None
            if stagiaire.encadrant_id:
                encadrant = db.query(Encadrant).filter(
                    Encadrant.id == stagiaire.encadrant_id
                ).first()

            departement = None
            if stagiaire.departement_id:
                departement = db.query(Departement).filter(
                    Departement.id == stagiaire.departement_id
                ).first()

            formations = db.query(Formation).filter(
                Formation.stagiaire_id == stagiaire.id
            ).order_by(Formation.ordre).all()

            competences = db.query(Competence).filter(
                Competence.stagiaire_id == stagiaire.id
            ).all()

            profil["telephone"] = stagiaire.telephone
            profil["localisation"] = stagiaire.localisation
            profil["photo_url"] = stagiaire.photo_url
            profil["ecole"] = stagiaire.ecole if hasattr(stagiaire, "ecole") else None
            profil["niveau_etudes"] = stagiaire.niveau_etudes
            profil["specialisation"] = stagiaire.specialisation
            profil["formations"] = [
                {
                    "id": f.id,
                    "etablissement": f.etablissement,
                    "diplome": f.diplome,
                    "date_debut": f.date_debut,
                    "date_fin": f.date_fin,
                }
                for f in formations
            ]
            profil["competences"] = [
                {"id": c.id, "nom": c.nom} for c in competences
            ]

            profil["stage"] = {
                "prenom": stagiaire.prenom,
                "nom": stagiaire.nom,
                "type_stage": stagiaire.type_stage,
                "date_debut": stagiaire.date_debut,
                "date_fin": stagiaire.date_fin,
                "statut": stagiaire.statut,
                "departement": departement.nom if departement is not None else None,
                "encadrant": {
                    "nom": encadrant.nom,
                    "prenom": encadrant.prenom,
                    "email": encadrant.email,
                } if encadrant is not None else None,
            }

    return profil


@app.put("/moi/mot-de-passe")
def changer_mon_mot_de_passe(
    donnees: ChangementMotDePasse,
    db: Session = Depends(get_db),
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
):
    if not verifier_mot_de_passe(donnees.ancien_mot_de_passe, utilisateur_courant.mot_de_passe_hash):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Ancien mot de passe incorrect",
        )
    utilisateur_courant.mot_de_passe_hash = hasher_mot_de_passe(donnees.nouveau_mot_de_passe)
    db.commit()
    return {"message": "Mot de passe modifié avec succès"}

@app.post("/mot-de-passe-oublie")
def mot_de_passe_oublie(donnees: MotDePasseOublie, db: Session = Depends(get_db)):
    utilisateur = db.query(Utilisateur).filter(Utilisateur.email == donnees.email).first()
    if utilisateur is None:
        return {"message": "Si cet email existe, un lien a ete envoye"}
    token = secrets.token_urlsafe(32)
    tokens_reinitialisation[token] = utilisateur.email
    print(f"[SIMULATION] Token pour {utilisateur.email} : {token}")
    return {"message": "Si cet email existe, un lien a ete envoye"}


@app.put("/moi")
def modifier_mon_profil(
    profil: ProfilUpdate,
    db: Session = Depends(get_db),
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
):
    # Mise à jour du nom si fourni
    if profil.nom is not None:
        utilisateur_courant.nom = profil.nom
    
    # Mise à jour de l'email si fourni
    if profil.email is not None:
        # Vérifier que le nouvel email n'est pas déjà utilisé par un autre utilisateur
        existe = db.query(Utilisateur).filter(
            Utilisateur.email == profil.email,
            Utilisateur.id != utilisateur_courant.id
        ).first()
        if existe:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Cet email est déjà utilisé par un autre compte",
            )
        utilisateur_courant.email = profil.email

    # Mise à jour des champs spécifiques au stagiaire (téléphone, localisation)
    if utilisateur_courant.role == "stagiaire" and utilisateur_courant.stagiaire_id:
        stagiaire = db.query(Stagiaire).filter(
            Stagiaire.id == utilisateur_courant.stagiaire_id
        ).first()
        if stagiaire is not None:
            if profil.telephone is not None:
                stagiaire.telephone = profil.telephone
            if profil.localisation is not None:
                stagiaire.localisation = profil.localisation

    db.commit()
    db.refresh(utilisateur_courant)
    
    return {
        "message": "Profil mis à jour avec succès",
        "utilisateur": {
            "id": utilisateur_courant.id,
            "nom": utilisateur_courant.nom,
            "email": utilisateur_courant.email,
            "role": utilisateur_courant.role,
        },
    }

# ============================================================
# Stagiaires (table principale)
# ============================================================




@app.post("/stagiaires")
def ajouter_stagiaire(
    stagiaire: StagiaireCreate,
    db: Session = Depends(get_db),
    utilisateur_courant: Utilisateur = Depends(exiger_role("admin_rh")),
):
    nouveau = Stagiaire(**stagiaire.dict())
    db.add(nouveau)
    db.commit()
    db.refresh(nouveau)
    return nouveau


@app.put("/stagiaires/{stagiaire_id}")
def modifier_stagiaire(
    stagiaire_id: int,
    stagiaire_maj: StagiaireCreate,
    db: Session = Depends(get_db),
    utilisateur_courant: Utilisateur = Depends(exiger_role("admin_rh")),
):
    stagiaire = db.query(Stagiaire).filter(Stagiaire.id == stagiaire_id).first()
    if stagiaire is None:
        return {"erreur": "Stagiaire non trouvé"}
    for champ, valeur in stagiaire_maj.dict().items():
        setattr(stagiaire, champ, valeur)
    db.commit()
    db.refresh(stagiaire)
    return stagiaire


@app.delete("/stagiaires/{stagiaire_id}")
def supprimer_stagiaire(
    stagiaire_id: int,
    db: Session = Depends(get_db),
    utilisateur_courant: Utilisateur = Depends(exiger_role("admin_rh")),
):
    obj = db.query(Stagiaire).filter(Stagiaire.id == stagiaire_id).first()
    if obj is None:
        return {"erreur": "Stagiaire non trouvé"}
    db.delete(obj)
    db.commit()
    return {"message": f"Stagiaire {stagiaire_id} supprimé"}


# ============================================================
# Documents
# ============================================================

@app.get("/documents")
def liste_documents(
    db: Session = Depends(get_db),
    utilisateur_courant: Utilisateur = Depends(exiger_role("rh", "admin_rh")),
):
    return db.query(Document).all()


@app.post("/documents")
def ajouter_document(
    document: DocumentCreate,
    db: Session = Depends(get_db),
    utilisateur_courant: Utilisateur = Depends(exiger_role("rh", "admin_rh")),
):
    nouveau = Document(**document.dict())
    db.add(nouveau)
    db.commit()
    db.refresh(nouveau)
    return nouveau


@app.delete("/documents/{document_id}")
def supprimer_document(
    document_id: int,
    db: Session = Depends(get_db),
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
):
    obj = db.query(Document).filter(Document.id == document_id).first()
    if obj is None:
        return {"erreur": "Document non trouvé"}

    est_admin = utilisateur_courant.role == "admin_rh"
    est_proprietaire = (
        utilisateur_courant.role == "stagiaire"
        and utilisateur_courant.stagiaire_id == obj.stagiaire_id
    )
    if not (est_admin or est_proprietaire):
        raise HTTPException(status_code=403, detail="Non autorise a supprimer ce document")

    db.delete(obj)
    db.commit()
    return {"message": f"Document {document_id} supprimé"}


class StagiaireResumeOut(BaseModel):
    id: int
    nom: str
    prenom: str

    class Config:
        from_attributes = True


class DocumentRHOut(DocumentOut):
    stagiaire: StagiaireResumeOut | None = None

    @classmethod
    def depuis_document_rh(cls, doc):
        base = DocumentOut.depuis_document(doc)
        stagiaire_out = None
        if doc.stagiaire_id:
            stagiaire_obj = getattr(doc, "_stagiaire_charge", None)
            if stagiaire_obj:
                stagiaire_out = StagiaireResumeOut.from_orm(stagiaire_obj)
        return cls(**base.dict(), stagiaire=stagiaire_out)


class DocumentStatutUpdate(BaseModel):
    statut: str
    commentaire: str | None = None


@app.get("/rh/documents/export")
def exporter_documents_excel(
    db: Session = Depends(get_db),
    utilisateur_courant: Utilisateur = Depends(exiger_role("rh", "admin_rh")),
):
    stagiaires = db.query(Stagiaire).all()
    departements = {d.id: d.nom for d in db.query(Departement).all()}
    encadrants = {e.id: f"{e.prenom} {e.nom}" for e in db.query(Encadrant).all()}
    documents = db.query(Document).all()

    docs_par_stagiaire = {}
    for doc in documents:
        docs_par_stagiaire.setdefault(doc.stagiaire_id, []).append(doc)

    wb = Workbook()
    ws = wb.active
    ws.title = "Documents"

    entetes = [
        "Stagiaire", "CIN", "Departement", "Encadrant",
        "Type de document", "Statut", "Date du document",
        "Commentaire", "Origine",
    ]
    ws.append(entetes)

    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for stagiaire in stagiaires:
        nom_complet = f"{stagiaire.prenom} {stagiaire.nom}"
        dept_nom = departements.get(stagiaire.departement_id, "")
        enc_nom = encadrants.get(stagiaire.encadrant_id, "")
        docs = docs_par_stagiaire.get(stagiaire.id, [])

        if not docs:
            ws.append([nom_complet, stagiaire.cin or "", dept_nom, enc_nom, "", "Aucun document", "", "", ""])
            continue

        for doc in docs:
            ws.append([
                nom_complet,
                stagiaire.cin or "",
                dept_nom,
                enc_nom,
                doc.type_document or "",
                doc.statut or "",
                doc.date_document.strftime("%d/%m/%Y") if doc.date_document else "",
                doc.commentaire or "",
                doc.origine or "",
            ])

    largeurs = [22, 14, 16, 20, 20, 14, 16, 30, 12]
    for i, largeur in enumerate(largeurs, start=1):
        ws.column_dimensions[chr(64 + i)].width = largeur

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nom_fichier = f"documents_stagiaires_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nom_fichier}"'},
    )


@app.get("/rh/documents", response_model=list[DocumentRHOut])
def liste_documents_rh(
    statut: str | None = None,
    db: Session = Depends(get_db),
    utilisateur_courant: Utilisateur = Depends(exiger_role("rh", "admin_rh")),
):
    requete = db.query(Document)
    if statut:
        requete = requete.filter(Document.statut == statut)
    documents = requete.all()

    resultats = []
    for doc in documents:
        stagiaire_obj = db.query(Stagiaire).filter(Stagiaire.id == doc.stagiaire_id).first()
        doc._stagiaire_charge = stagiaire_obj
        resultats.append(DocumentRHOut.depuis_document_rh(doc))
    return resultats


@app.patch("/documents/{document_id}/statut", response_model=DocumentOut)
def modifier_statut_document(
    document_id: int,
    donnees: DocumentStatutUpdate,
    db: Session = Depends(get_db),
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
):
    statuts_autorises = {"en_attente", "valide", "genere", "recu", "manquant", "refuse", "a_completer"}
    if donnees.statut not in statuts_autorises:
        raise HTTPException(status_code=400, detail="Statut invalide")

    doc = db.query(Document).filter(Document.id == document_id).first()
    if doc is None:
        raise HTTPException(status_code=404, detail="Document introuvable")

    if utilisateur_courant.role == "encadrant":
        verifier_stagiaire_assigne(doc.stagiaire_id, utilisateur_courant, db)
    elif utilisateur_courant.role not in ("rh", "admin_rh"):
        raise HTTPException(status_code=403, detail="Non autorise a modifier le statut")

    doc.statut = donnees.statut
    doc.commentaire = donnees.commentaire
    db.commit()
    db.refresh(doc)

    if donnees.statut == "a_completer" and doc.stagiaire_id:
        notification_stagiaire = Notification(
            stagiaire_id=doc.stagiaire_id,
            categorie="document",
            urgence="normale",
            titre="Document a completer",
            contenu=donnees.commentaire or "Un de vos documents necessite une action de votre part.",
        )
        db.add(notification_stagiaire)
        db.commit()

    return DocumentOut.depuis_document(doc)


TYPES_DOCUMENTS_AUTORISES = {
    "convention", "attestation", "rapport_intermediaire", "rapport_final",
    "lettre_affectation", "badge_photo", "fiche_securite", "certificat",
}
EXTENSIONS_AUTORISEES = {".pdf", ".jpg", ".jpeg", ".png", ".doc", ".docx"}
TAILLE_MAX_OCTETS = 10 * 1024 * 1024  # 10 Mo


@app.post("/documents/upload")
async def uploader_document(
    type_document: str = Form(...),
    fichier: UploadFile = File(...),
    db: Session = Depends(get_db),
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
):
    if utilisateur_courant.role != "stagiaire" or not utilisateur_courant.stagiaire_id:
        raise HTTPException(status_code=403, detail="Réservé aux stagiaires")

    if type_document not in TYPES_DOCUMENTS_AUTORISES:
        raise HTTPException(status_code=400, detail="Type de document invalide")

    extension = os.path.splitext(fichier.filename)[1].lower()
    if extension not in EXTENSIONS_AUTORISEES:
        raise HTTPException(status_code=400, detail="Type de fichier non autorisé")

    contenu = await fichier.read()
    if len(contenu) > TAILLE_MAX_OCTETS:
        raise HTTPException(status_code=400, detail="Fichier trop volumineux (max 10 Mo)")

    nom_unique = f"{uuid.uuid4().hex}{extension}"
    chemin_disque = os.path.join(UPLOAD_DIR, nom_unique)
    with open(chemin_disque, "wb") as f:
        f.write(contenu)

    nouveau_document = Document(
        stagiaire_id=utilisateur_courant.stagiaire_id,
        type_document=type_document,
        statut="en_attente",
        date_document=date.today(),
        fichier_url=f"/uploads/{nom_unique}",
        taille_octets=len(contenu),
        origine="stagiaire",
    )
    db.add(nouveau_document)
    db.commit()
    db.refresh(nouveau_document)


    stagiaire_concerne = db.query(Stagiaire).filter(
        Stagiaire.id == utilisateur_courant.stagiaire_id
    ).first()
    if stagiaire_concerne and stagiaire_concerne.encadrant_id:
        notification_encadrant = Notification(
            encadrant_id=stagiaire_concerne.encadrant_id,
            stagiaire_id=stagiaire_concerne.id,
            categorie="document",
            urgence="normale",
            titre="Nouveau document déposé",
            contenu=f"{stagiaire_concerne.nom} a déposé un document ({type_document}).",
        )
        db.add(notification_encadrant)
        db.commit()

    return DocumentOut.depuis_document(nouveau_document)


@app.post("/documents/upload-rh")
async def uploader_document_rh(
    stagiaire_id: int = Form(...),
    type_document: str = Form(...),
    fichier: UploadFile = File(...),
    date_document: str = Form(None),
    db: Session = Depends(get_db),
    utilisateur_courant: Utilisateur = Depends(exiger_role("rh", "admin_rh")),
):
    stagiaire_cible = db.query(Stagiaire).filter(Stagiaire.id == stagiaire_id).first()
    if not stagiaire_cible:
        raise HTTPException(status_code=404, detail="Stagiaire introuvable")

    extension = os.path.splitext(fichier.filename)[1].lower()
    if extension not in EXTENSIONS_AUTORISEES:
        raise HTTPException(status_code=400, detail="Type de fichier non autorisé")

    contenu = await fichier.read()
    if len(contenu) > TAILLE_MAX_OCTETS:
        raise HTTPException(status_code=400, detail="Fichier trop volumineux (max 10 Mo)")

    nom_unique = f"{uuid.uuid4().hex}{extension}"
    chemin_disque = os.path.join(UPLOAD_DIR, nom_unique)
    with open(chemin_disque, "wb") as f:
        f.write(contenu)

    date_finale = date.today()
    if date_document:
        try:
            date_finale = date.fromisoformat(date_document)
        except ValueError:
            pass

    nouveau_document = Document(
        stagiaire_id=stagiaire_id,
        type_document=type_document,
        statut="valide",
        date_document=date_finale,
        fichier_url=f"/uploads/{nom_unique}",
        taille_octets=len(contenu),
        origine="rh",
    )
    db.add(nouveau_document)
    db.commit()
    db.refresh(nouveau_document)

    notification_stagiaire = Notification(
        stagiaire_id=stagiaire_id,
        categorie="document",
        urgence="normale",
        titre="Nouveau document disponible",
        contenu=f"Le service RH vous a transmis un document ({type_document}).",
    )
    db.add(notification_stagiaire)
    db.commit()

    return DocumentOut.depuis_document(nouveau_document)


@app.put("/documents/{document_id}")
async def modifier_document(
    document_id: int,
    type_document: str = Form(None),
    fichier: UploadFile = File(None),
    db: Session = Depends(get_db),
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
):
    if utilisateur_courant.role != "stagiaire" or not utilisateur_courant.stagiaire_id:
        raise HTTPException(status_code=403, detail="Reserve aux stagiaires")

    document = db.query(Document).filter(
        Document.id == document_id,
        Document.stagiaire_id == utilisateur_courant.stagiaire_id,
    ).first()
    if not document:
        raise HTTPException(status_code=404, detail="Document introuvable")

    if document.statut != "en_attente":
        raise HTTPException(status_code=400, detail="Seuls les documents en attente peuvent etre modifies")

    if type_document:
        if type_document not in TYPES_DOCUMENTS_AUTORISES:
            raise HTTPException(status_code=400, detail="Type de document invalide")
        document.type_document = type_document

    if fichier:
        extension = os.path.splitext(fichier.filename)[1].lower()
        if extension not in EXTENSIONS_AUTORISEES:
            raise HTTPException(status_code=400, detail="Type de fichier non autorise")
        contenu = await fichier.read()
        if len(contenu) > TAILLE_MAX_OCTETS:
            raise HTTPException(status_code=400, detail="Fichier trop volumineux (max 10 Mo)")
        nom_unique = f"{uuid.uuid4().hex}{extension}"
        chemin_disque = os.path.join(UPLOAD_DIR, nom_unique)
        with open(chemin_disque, "wb") as f:
            f.write(contenu)
        document.fichier_url = f"/uploads/{nom_unique}"
        document.taille_octets = len(contenu)
        document.date_document = date.today()

    db.commit()
    db.refresh(document)
    return DocumentOut.depuis_document(document)


EXTENSIONS_PHOTO_AUTORISEES = {".jpg", ".jpeg", ".png", ".webp"}
TAILLE_MAX_PHOTO_OCTETS = 5 * 1024 * 1024  # 5 Mo


@app.post("/moi/photo")
async def uploader_photo_profil(
    fichier: UploadFile = File(...),
    db: Session = Depends(get_db),
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
):
    if utilisateur_courant.role != "stagiaire" or not utilisateur_courant.stagiaire_id:
        raise HTTPException(status_code=403, detail="Réservé aux stagiaires")

    extension = os.path.splitext(fichier.filename)[1].lower()
    if extension not in EXTENSIONS_PHOTO_AUTORISEES:
        raise HTTPException(status_code=400, detail="Format d'image non autorisé (jpg, jpeg, png, webp uniquement)")

    contenu = await fichier.read()
    if len(contenu) > TAILLE_MAX_PHOTO_OCTETS:
        raise HTTPException(status_code=400, detail="Image trop volumineuse (max 5 Mo)")

    stagiaire = db.query(Stagiaire).filter(Stagiaire.id == utilisateur_courant.stagiaire_id).first()
    if not stagiaire:
        raise HTTPException(status_code=404, detail="Stagiaire introuvable")

    nom_unique = f"{uuid.uuid4().hex}{extension}"
    chemin_disque = os.path.join(UPLOAD_DIR, nom_unique)
    with open(chemin_disque, "wb") as f:
        f.write(contenu)

    stagiaire.photo_url = f"/uploads/{nom_unique}"
    db.commit()
    db.refresh(stagiaire)

    return {"photo_url": stagiaire.photo_url}


# ============================================================
# Présences
# ============================================================

@app.get("/presences")
def liste_presences(
    db: Session = Depends(get_db),
    utilisateur_courant: Utilisateur = Depends(exiger_role("rh", "admin_rh")),
):
    return db.query(Presence).all()


@app.post("/presences")
def ajouter_presence(
    presence: PresenceCreate,
    db: Session = Depends(get_db),
    utilisateur_courant: Utilisateur = Depends(exiger_role("admin_rh")),
):
    nouvelle = Presence(**presence.dict())
    db.add(nouvelle)
    db.commit()
    db.refresh(nouvelle)
    return nouvelle


@app.put("/presences/{presence_id}")
def modifier_presence(
    presence_id: int,
    presence: PresenceCreate,
    db: Session = Depends(get_db),
    utilisateur_courant: Utilisateur = Depends(exiger_role("admin_rh")),
):
    obj = db.query(Presence).filter(Presence.id == presence_id).first()
    if obj is None:
        raise HTTPException(status_code=404, detail="Presence non trouvee")
    for champ, valeur in presence.dict().items():
        setattr(obj, champ, valeur)
    db.commit()
    db.refresh(obj)
    return obj


@app.delete("/presences/{presence_id}")
def supprimer_presence(
    presence_id: int,
    db: Session = Depends(get_db),
    utilisateur_courant: Utilisateur = Depends(exiger_role("admin_rh")),
):
    obj = db.query(Presence).filter(Presence.id == presence_id).first()
    if obj is None:
        return {"erreur": "Présence non trouvée"}
    db.delete(obj)
    db.commit()
    return {"message": f"Présence {presence_id} supprimée"}


@app.put("/moi/presences/{presence_id}")
def modifier_ma_presence(
    presence_id: int,
    donnees: PresenceStagiaireUpdate,
    db: Session = Depends(get_db),
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
):
    if utilisateur_courant.role != "stagiaire" or not utilisateur_courant.stagiaire_id:
        raise HTTPException(status_code=403, detail="Reserve aux stagiaires")

    presence = db.query(Presence).filter(
        Presence.id == presence_id,
        Presence.stagiaire_id == utilisateur_courant.stagiaire_id,
    ).first()
    if not presence:
        raise HTTPException(status_code=404, detail="Presence introuvable")

    if donnees.heure_arrivee is not None:
        presence.heure_arrivee = donnees.heure_arrivee
    if donnees.heure_depart is not None:
        presence.heure_depart = donnees.heure_depart

    db.commit()
    db.refresh(presence)
    return presence


# ============================================================
# Activités
# ============================================================

@app.get("/activites")
def liste_activites(db: Session = Depends(get_db)):
    return db.query(Activite).all()


@app.post("/activites")
def ajouter_activite(
    activite: ActiviteCreate,
    db: Session = Depends(get_db),
    utilisateur_courant: Utilisateur = Depends(exiger_role("admin_rh")),
):
    nouvelle = Activite(**activite.dict())
    db.add(nouvelle)
    db.commit()
    db.refresh(nouvelle)
    return nouvelle


@app.delete("/activites/{activite_id}")
def supprimer_activite(
    activite_id: int,
    db: Session = Depends(get_db),
    utilisateur_courant: Utilisateur = Depends(exiger_role("admin_rh")),
):
    obj = db.query(Activite).filter(Activite.id == activite_id).first()
    if obj is None:
        return {"erreur": "Activité non trouvée"}
    db.delete(obj)
    db.commit()
    return {"message": f"Activité {activite_id} supprimée"}



@app.post("/reinitialiser-mot-de-passe")
def reinitialiser_mot_de_passe(donnees: ReinitialiserMotDePasse, db: Session = Depends(get_db)):
    email = tokens_reinitialisation.get(donnees.token)
    if email is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Token invalide ou expire")
    utilisateur = db.query(Utilisateur).filter(Utilisateur.email == email).first()
    if utilisateur is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Utilisateur introuvable")
    utilisateur.mot_de_passe_hash = hasher_mot_de_passe(donnees.nouveau_mot_de_passe)
    db.commit()
    del tokens_reinitialisation[donnees.token]
    return {"message": "Mot de passe reinitialise avec succes"}
@app.get("/moi/documents", response_model=list[DocumentOut])
def obtenir_mes_documents(
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "stagiaire" or not utilisateur_courant.stagiaire_id:
        raise HTTPException(status_code=403, detail="Reserve aux stagiaires")

    documents = (
        db.query(Document)
        .filter(Document.stagiaire_id == utilisateur_courant.stagiaire_id)
        .order_by(Document.date_document.desc())
        .all()
    )
    return [DocumentOut.depuis_document(doc) for doc in documents]


@app.get("/moi/documents/stats")
def obtenir_stats_documents(
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "stagiaire" or not utilisateur_courant.stagiaire_id:
        raise HTTPException(status_code=403, detail="Reserve aux stagiaires")

    documents = (
        db.query(Document)
        .filter(Document.stagiaire_id == utilisateur_courant.stagiaire_id)
        .filter(Document.fichier_url.isnot(None))
        .all()
    )
    statuts_valides = {"valide", "genere", "recu"}
    total = len(documents)
    valides = sum(1 for d in documents if d.statut in statuts_valides)
    en_attente = sum(1 for d in documents if d.statut == "en_attente")
    refuses = sum(1 for d in documents if d.statut == "refuse")
    return {
        "total": total,
        "valides": valides,
        "en_attente": en_attente,
        "refuses": refuses,
    }


# ============================================================
# Compétences techniques du stagiaire connecté
# ============================================================

@app.post("/moi/competences", response_model=CompetenceOut)
def ajouter_ma_competence(
    donnees: CompetenceCreate,
    db: Session = Depends(get_db),
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
):
    if utilisateur_courant.role != "stagiaire" or not utilisateur_courant.stagiaire_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Seul un stagiaire peut ajouter une compétence",
        )

    nom_propre = donnees.nom.strip()
    if not nom_propre:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Le nom de la compétence ne peut pas être vide",
        )

    existe = db.query(Competence).filter(
        Competence.stagiaire_id == utilisateur_courant.stagiaire_id,
        Competence.nom.ilike(nom_propre),
    ).first()
    if existe:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Cette compétence est déjà dans votre liste",
        )

    competence = Competence(stagiaire_id=utilisateur_courant.stagiaire_id, nom=nom_propre)
    db.add(competence)
    db.commit()
    db.refresh(competence)
    return competence


@app.delete("/moi/competences/{competence_id}")
def supprimer_ma_competence(
    competence_id: int,
    db: Session = Depends(get_db),
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
):
    competence = db.query(Competence).filter(
        Competence.id == competence_id,
        Competence.stagiaire_id == utilisateur_courant.stagiaire_id,
    ).first()

    if competence is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Compétence introuvable",
        )

    db.delete(competence)
    db.commit()
    return {"message": "Compétence supprimée"}


# ============================================================
# Historique de formation du stagiaire connecté
# ============================================================

@app.post("/moi/formations", response_model=FormationOut)
def ajouter_ma_formation(
    donnees: FormationCreate,
    db: Session = Depends(get_db),
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
):
    if utilisateur_courant.role != "stagiaire" or not utilisateur_courant.stagiaire_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Seul un stagiaire peut ajouter une formation",
        )

    etablissement_propre = donnees.etablissement.strip()
    if not etablissement_propre:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Le nom de l'établissement ne peut pas être vide",
        )

    def parser_date(valeur):
        if not valeur:
            return None
        try:
            return date.fromisoformat(valeur)
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Format de date invalide, attendu AAAA-MM-JJ",
            )

    # On place la nouvelle formation en fin de liste (ordre croissant)
    ordre_suivant = db.query(Formation).filter(
        Formation.stagiaire_id == utilisateur_courant.stagiaire_id
    ).count()

    formation = Formation(
        stagiaire_id=utilisateur_courant.stagiaire_id,
        etablissement=etablissement_propre,
        diplome=donnees.diplome,
        date_debut=parser_date(donnees.date_debut),
        date_fin=parser_date(donnees.date_fin),
        ordre=ordre_suivant,
    )
    db.add(formation)
    db.commit()
    db.refresh(formation)
    return formation


@app.delete("/moi/formations/{formation_id}")
def supprimer_ma_formation(
    formation_id: int,
    db: Session = Depends(get_db),
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
):
    formation = db.query(Formation).filter(
        Formation.id == formation_id,
        Formation.stagiaire_id == utilisateur_courant.stagiaire_id,
    ).first()

    if formation is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Formation introuvable",
        )

    db.delete(formation)
    db.commit()
    return {"message": "Formation supprimée"}


class ActiviteStagiaireCreate(BaseModel):
    action: str
    priorite: str | None = "moyenne"
    echeance: date | None = None


class ActiviteStagiaireUpdate(BaseModel):
    statut: str | None = None
    progression: int | None = None
    priorite: str | None = None
    echeance: date | None = None


@app.get("/moi/activites")
def obtenir_mes_activites(
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "stagiaire" or not utilisateur_courant.stagiaire_id:
        raise HTTPException(status_code=403, detail="Reserve aux stagiaires")

    activites = (
        db.query(Activite)
        .filter(Activite.stagiaire_id == utilisateur_courant.stagiaire_id)
        .order_by(Activite.date_action.desc())
        .all()
    )
    return activites


@app.post("/moi/activites")
def creer_mon_activite(
    donnees: ActiviteStagiaireCreate,
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "stagiaire" or not utilisateur_courant.stagiaire_id:
        raise HTTPException(status_code=403, detail="Reserve aux stagiaires")

    nouvelle = Activite(
        stagiaire_id=utilisateur_courant.stagiaire_id,
        action=donnees.action,
        statut="a_faire",
        priorite=donnees.priorite or "moyenne",
        echeance=donnees.echeance,
        progression=0,
    )
    db.add(nouvelle)
    db.commit()
    db.refresh(nouvelle)
    return nouvelle


@app.put("/moi/activites/{activite_id}")
def modifier_mon_activite(
    activite_id: int,
    donnees: ActiviteStagiaireUpdate,
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    activite = db.query(Activite).filter(
        Activite.id == activite_id,
        Activite.stagiaire_id == utilisateur_courant.stagiaire_id,
    ).first()

    if activite is None:
        raise HTTPException(status_code=404, detail="Activite introuvable")

    if donnees.statut is not None:
        activite.statut = donnees.statut
    if donnees.progression is not None:
        activite.progression = donnees.progression
    if donnees.priorite is not None:
        activite.priorite = donnees.priorite
    if donnees.echeance is not None:
        activite.echeance = donnees.echeance

    db.commit()
    db.refresh(activite)
    return activite


@app.delete("/moi/activites/{activite_id}")
def supprimer_mon_activite(
    activite_id: int,
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    activite = db.query(Activite).filter(
        Activite.id == activite_id,
        Activite.stagiaire_id == utilisateur_courant.stagiaire_id,
    ).first()

    if activite is None:
        raise HTTPException(status_code=404, detail="Activite introuvable")

    db.delete(activite)
    db.commit()
    return {"message": "Activite supprimee"}


class SessionFormationCreate(BaseModel):
    titre: str
    date_session: date
    heure: str | None = None
    salle: str | None = None
    description: str | None = None


@app.get("/sessions")
def liste_sessions(db: Session = Depends(get_db)):
    return db.query(SessionFormation).order_by(SessionFormation.date_session).all()


@app.get("/sessions/a-venir")
def sessions_a_venir(db: Session = Depends(get_db)):
    aujourdhui = date.today()
    return (
        db.query(SessionFormation)
        .filter(SessionFormation.date_session >= aujourdhui)
        .order_by(SessionFormation.date_session)
        .all()
    )


@app.post("/sessions")
def ajouter_session(
    session_donnees: SessionFormationCreate,
    db: Session = Depends(get_db),
    utilisateur_courant: Utilisateur = Depends(exiger_role("admin_rh")),
):
    nouvelle = SessionFormation(**session_donnees.dict())
    db.add(nouvelle)
    db.commit()
    db.refresh(nouvelle)
    return nouvelle


@app.put("/sessions/{session_id}")
def modifier_session(
    session_id: int,
    session_donnees: SessionFormationCreate,
    db: Session = Depends(get_db),
    utilisateur_courant: Utilisateur = Depends(exiger_role("admin_rh")),
):
    session_obj = db.query(SessionFormation).filter(SessionFormation.id == session_id).first()
    if session_obj is None:
        raise HTTPException(status_code=404, detail="Session non trouvee")
    for champ, valeur in session_donnees.dict().items():
        setattr(session_obj, champ, valeur)
    db.commit()
    db.refresh(session_obj)
    return session_obj


@app.delete("/sessions/{session_id}")
def supprimer_session(
    session_id: int,
    db: Session = Depends(get_db),
    utilisateur_courant: Utilisateur = Depends(exiger_role("admin_rh")),
):
    obj = db.query(SessionFormation).filter(SessionFormation.id == session_id).first()
    if obj is None:
        raise HTTPException(status_code=404, detail="Session non trouvee")
    db.delete(obj)
    db.commit()
    return {"message": f"Session {session_id} supprimee"}


@app.get("/presences/jour")
def presences_du_jour(
    db: Session = Depends(get_db),
    utilisateur_courant: Utilisateur = Depends(exiger_role("rh", "admin_rh")),
):
    aujourdhui = date.today()
    total = db.query(Stagiaire).filter(Stagiaire.statut == "en_cours").count()
    present = (
        db.query(Presence)
        .filter(Presence.date == aujourdhui, Presence.present == True)
        .count()
    )
    return {"present": present, "total": total}


@app.get("/documents/manquants")
def documents_manquants(
    db: Session = Depends(get_db),
    utilisateur_courant: Utilisateur = Depends(exiger_role("admin_rh")),
):
    count = (
        db.query(Document)
        .join(Stagiaire, Document.stagiaire_id == Stagiaire.id)
        .filter(Stagiaire.statut == "en_cours")
        .filter(Document.statut.in_(["en_attente", "a_reviser"]))
        .count()
    )
    return {"documents_manquants": count}



@app.get("/moi/presences")
def obtenir_mes_presences(
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "stagiaire" or not utilisateur_courant.stagiaire_id:
        raise HTTPException(status_code=403, detail="Reserve aux stagiaires")

    presences = (
        db.query(Presence)
        .filter(Presence.stagiaire_id == utilisateur_courant.stagiaire_id)
        .order_by(Presence.date.desc())
        .all()
    )
    return presences


@app.get("/mes-stagiaires/confirmations-evenement/{evenement_cle}")
def confirmations_evenement_de_mes_stagiaires(
    evenement_cle: str,
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "encadrant" or not utilisateur_courant.encadrant_id:
        raise HTTPException(status_code=403, detail="Reserve aux encadrants")

    mes_stagiaires = (
        db.query(Stagiaire)
        .filter(Stagiaire.encadrant_id == utilisateur_courant.encadrant_id)
        .all()
    )

    confirmations = (
        db.query(ConfirmationEvenement)
        .filter(
            ConfirmationEvenement.evenement_cle == evenement_cle,
            ConfirmationEvenement.stagiaire_id.in_([s.id for s in mes_stagiaires]),
        )
        .all()
    )
    ids_confirmes = {c.stagiaire_id: c.date_confirmation for c in confirmations}

    resultat = []
    for s in mes_stagiaires:
        resultat.append({
            "stagiaire_id": s.id,
            "nom": f"{s.prenom} {s.nom}",
            "confirme": s.id in ids_confirmes,
            "date_confirmation": ids_confirmes.get(s.id),
        })

    resultat.sort(key=lambda x: (not x["confirme"], x["nom"]))
    return resultat


@app.get("/moi/confirmer-evenement/{evenement_cle}")
def verifier_confirmation_evenement(
    evenement_cle: str,
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "stagiaire" or not utilisateur_courant.stagiaire_id:
        raise HTTPException(status_code=403, detail="Reserve aux stagiaires")

    confirmation = (
        db.query(ConfirmationEvenement)
        .filter(
            ConfirmationEvenement.stagiaire_id == utilisateur_courant.stagiaire_id,
            ConfirmationEvenement.evenement_cle == evenement_cle,
        )
        .first()
    )
    return {"confirme": confirmation is not None}


@app.post("/moi/confirmer-evenement/{evenement_cle}")
def confirmer_presence_evenement(
    evenement_cle: str,
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "stagiaire" or not utilisateur_courant.stagiaire_id:
        raise HTTPException(status_code=403, detail="Reserve aux stagiaires")

    existe = (
        db.query(ConfirmationEvenement)
        .filter(
            ConfirmationEvenement.stagiaire_id == utilisateur_courant.stagiaire_id,
            ConfirmationEvenement.evenement_cle == evenement_cle,
        )
        .first()
    )
    if existe:
        return {"message": "Deja confirme", "confirme": True}

    nouvelle = ConfirmationEvenement(
        stagiaire_id=utilisateur_courant.stagiaire_id,
        evenement_cle=evenement_cle,
    )
    db.add(nouvelle)
    db.commit()
    return {"message": "Presence confirmee", "confirme": True}


@app.post("/moi/presences/pointer")
def pointer_ma_presence(
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "stagiaire" or not utilisateur_courant.stagiaire_id:
        raise HTTPException(status_code=403, detail="Reserve aux stagiaires")

    aujourdhui = date.today()
    heure_actuelle = datetime.now().strftime("%H:%M")

    presence_du_jour = (
        db.query(Presence)
        .filter(
            Presence.stagiaire_id == utilisateur_courant.stagiaire_id,
            Presence.date == aujourdhui,
        )
        .first()
    )

    if presence_du_jour is None:
        nouvelle = Presence(
            stagiaire_id=utilisateur_courant.stagiaire_id,
            date=aujourdhui,
            present=True,
            marque_par=utilisateur_courant.id,
            heure_arrivee=heure_actuelle,
            heure_depart=None,
        )
        db.add(nouvelle)
        db.commit()
        db.refresh(nouvelle)
        return {"message": "Arrivee enregistree", "presence": nouvelle}

    if presence_du_jour.heure_arrivee and not presence_du_jour.heure_depart:
        presence_du_jour.heure_depart = heure_actuelle
        db.commit()
        db.refresh(presence_du_jour)
        return {"message": "Depart enregistre", "presence": presence_du_jour}

    raise HTTPException(
        status_code=400,
        detail="Vous avez deja pointe votre arrivee et votre depart aujourd'hui",
    )


@app.get("/moi/presences/stats")
def obtenir_stats_presences(
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "stagiaire" or not utilisateur_courant.stagiaire_id:
        raise HTTPException(status_code=403, detail="Reserve aux stagiaires")

    aujourdhui = date.today()
    debut_mois = aujourdhui.replace(day=1)

    presences_mois = (
        db.query(Presence)
        .filter(
            Presence.stagiaire_id == utilisateur_courant.stagiaire_id,
            Presence.date >= debut_mois,
            Presence.date <= aujourdhui,
        )
        .all()
    )

    total_enregistrements = len(presences_mois)
    jours_presents = [p for p in presences_mois if p.present]
    jours_absents = [p for p in presences_mois if not p.present]

    taux_presence = (
        round(len(jours_presents) / total_enregistrements * 100, 1)
        if total_enregistrements > 0
        else 0
    )

    total_heures = 0.0
    for p in jours_presents:
        if p.heure_arrivee and p.heure_depart:
            try:
                h_arr, m_arr = map(int, p.heure_arrivee.split(":"))
                h_dep, m_dep = map(int, p.heure_depart.split(":"))
                duree = (h_dep + m_dep / 60) - (h_arr + m_arr / 60)
                if duree > 0:
                    total_heures += duree
            except (ValueError, AttributeError):
                pass

    dernier = (
        db.query(Presence)
        .filter(Presence.stagiaire_id == utilisateur_courant.stagiaire_id)
        .order_by(Presence.date.desc())
        .first()
    )
    statut_actuel = "present" if dernier and dernier.present else "absent"

    return {
        "total_heures_mois": round(total_heures, 1),
        "taux_presence": taux_presence,
        "jours_absence_mois": len(jours_absents),
        "statut_actuel": statut_actuel,
        "derniere_arrivee": dernier.heure_arrivee if dernier else None,
    }


@app.get("/moi/evaluations")
def obtenir_toutes_mes_evaluations(
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "encadrant" or not utilisateur_courant.encadrant_id:
        raise HTTPException(status_code=403, detail="Reserve aux encadrants")

    evaluations = (
        db.query(Evaluation)
        .filter(Evaluation.encadrant_id == utilisateur_courant.encadrant_id)
        .order_by(Evaluation.date_evaluation.desc())
        .all()
    )

    resultat = []
    for evaluation in evaluations:
        stagiaire = db.query(Stagiaire).filter(Stagiaire.id == evaluation.stagiaire_id).first()
        resultat.append({
            "id": evaluation.id,
            "titre": evaluation.titre,
            "note": evaluation.note,
            "date_evaluation": evaluation.date_evaluation,
            "statut": evaluation.statut,
            "commentaire_global": evaluation.commentaire_global,
            "stagiaire_id": evaluation.stagiaire_id,
            "stagiaire_nom": stagiaire.nom if stagiaire else None,
            "stagiaire_prenom": stagiaire.prenom if stagiaire else None,
            "stagiaire_photo_url": stagiaire.photo_url if stagiaire else None,
        })

    return resultat


@app.get("/moi/mes-stagiaires")
def obtenir_mes_stagiaires(
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "encadrant" or not utilisateur_courant.encadrant_id:
        raise HTTPException(status_code=403, detail="Reserve aux encadrants")

    stagiaires = (
        db.query(Stagiaire)
        .filter(Stagiaire.encadrant_id == utilisateur_courant.encadrant_id)
        .all()
    )

    resultat = []
    for s in stagiaires:
        resultat.append({
            "id": s.id,
            "nom": s.nom,
            "prenom": s.prenom,
            "email": s.email,
            "photo_url": s.photo_url,
            "ecole": s.ecole,
            "specialisation": s.specialisation,
            "type_stage": s.type_stage,
            "date_debut": s.date_debut,
            "date_fin": s.date_fin,
            "statut": s.statut,
        })

    return resultat


@app.get("/moi/evaluations-a-faire")
def obtenir_evaluations_a_faire(
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "encadrant" or not utilisateur_courant.encadrant_id:
        raise HTTPException(status_code=403, detail="Reserve aux encadrants")

    aujourdhui = date.today()
    limite = aujourdhui + timedelta(days=30)

    stagiaires = (
        db.query(Stagiaire)
        .filter(
            Stagiaire.encadrant_id == utilisateur_courant.encadrant_id,
            Stagiaire.date_fin >= aujourdhui,
            Stagiaire.date_fin <= limite,
        )
        .all()
    )

    resultat = []
    for s in stagiaires:
        deja_soumise = (
            db.query(Evaluation)
            .filter(Evaluation.stagiaire_id == s.id, Evaluation.statut == "soumise")
            .first()
        )
        if not deja_soumise:
            resultat.append({
                "id": s.id,
                "prenom": s.prenom,
                "nom": s.nom,
                "date_fin": s.date_fin,
            })

    return {"total": len(resultat), "stagiaires": resultat}


@app.get("/moi/stagiaires-disponibles", response_model=list[StagiaireDisponibleOut])
def obtenir_stagiaires_disponibles(
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "encadrant" or not utilisateur_courant.encadrant_id:
        raise HTTPException(status_code=403, detail="Reserve aux encadrants")

    stagiaires = (
        db.query(Stagiaire)
        .filter(Stagiaire.encadrant_id.is_(None))
        .all()
    )
    return stagiaires


@app.post("/moi/mes-stagiaires/{stagiaire_id}/affecter")
def affecter_stagiaire_a_moi(
    stagiaire_id: int,
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "encadrant" or not utilisateur_courant.encadrant_id:
        raise HTTPException(status_code=403, detail="Reserve aux encadrants")

    stagiaire = db.query(Stagiaire).filter(Stagiaire.id == stagiaire_id).first()
    if stagiaire is None:
        raise HTTPException(status_code=404, detail="Stagiaire introuvable")

    if stagiaire.encadrant_id is not None:
        raise HTTPException(status_code=409, detail="Ce stagiaire est deja affecte a un encadrant")

    stagiaire.encadrant_id = utilisateur_courant.encadrant_id
    db.commit()
    db.refresh(stagiaire)

    return {"id": stagiaire.id, "encadrant_id": stagiaire.encadrant_id}


@app.get("/moi/mes-stagiaires/presences")
def obtenir_presences_de_mes_stagiaires(
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "encadrant" or not utilisateur_courant.encadrant_id:
        raise HTTPException(status_code=403, detail="Reserve aux encadrants")

    mes_stagiaires_ids = [
        s.id
        for s in db.query(Stagiaire)
        .filter(Stagiaire.encadrant_id == utilisateur_courant.encadrant_id)
        .all()
    ]

    presences = (
        db.query(Presence)
        .filter(Presence.stagiaire_id.in_(mes_stagiaires_ids))
        .order_by(Presence.date.desc())
        .all()
    )
    return presences


@app.get("/moi/mes-stagiaires/{stagiaire_id}")
def obtenir_un_de_mes_stagiaires(
    stagiaire_id: int,
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "encadrant" or not utilisateur_courant.encadrant_id:
        raise HTTPException(status_code=403, detail="Reserve aux encadrants")

    stagiaire = db.query(Stagiaire).filter(Stagiaire.id == stagiaire_id).first()
    if stagiaire is None:
        raise HTTPException(status_code=404, detail="Stagiaire introuvable")

    if stagiaire.encadrant_id != utilisateur_courant.encadrant_id:
        raise HTTPException(status_code=403, detail="Ce stagiaire ne vous est pas assigne")

    etablissement = None
    if stagiaire.etablissement_id:
        etablissement = db.query(Etablissement).filter(Etablissement.id == stagiaire.etablissement_id).first()

    departement = None
    if stagiaire.departement_id:
        departement = db.query(Departement).filter(Departement.id == stagiaire.departement_id).first()

    return {
        "id": stagiaire.id,
        "nom": stagiaire.nom,
        "prenom": stagiaire.prenom,
        "email": stagiaire.email,
        "telephone": stagiaire.telephone,
        "cin": stagiaire.cin,
        "ecole": stagiaire.ecole,
        "etablissement": etablissement.nom if etablissement else None,
        "etablissement_id": stagiaire.etablissement_id,
        "niveau_etudes": stagiaire.niveau_etudes,
        "specialisation": stagiaire.specialisation,
        "type_stage": stagiaire.type_stage,
        "departement": departement.nom if departement else None,
        "departement_id": stagiaire.departement_id,
        "date_debut": stagiaire.date_debut,
        "date_fin": stagiaire.date_fin,
        "statut": stagiaire.statut,
        "localisation": stagiaire.localisation,
        "photo_url": stagiaire.photo_url,
    }


@app.get("/moi/mes-stagiaires/{stagiaire_id}/documents", response_model=list[DocumentOut])
def obtenir_documents_dun_de_mes_stagiaires(
    stagiaire_id: int,
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "encadrant" or not utilisateur_courant.encadrant_id:
        raise HTTPException(status_code=403, detail="Reserve aux encadrants")

    stagiaire = db.query(Stagiaire).filter(Stagiaire.id == stagiaire_id).first()
    if stagiaire is None:
        raise HTTPException(status_code=404, detail="Stagiaire introuvable")

    if stagiaire.encadrant_id != utilisateur_courant.encadrant_id:
        raise HTTPException(status_code=403, detail="Ce stagiaire ne vous est pas assigne")

    documents = (
        db.query(Document)
        .filter(Document.stagiaire_id == stagiaire_id)
        .order_by(Document.date_document.desc())
        .all()
    )
    return [DocumentOut.depuis_document(doc) for doc in documents]


def verifier_stagiaire_assigne(stagiaire_id: int, utilisateur_courant, db: Session):
    if utilisateur_courant.role != "encadrant" or not utilisateur_courant.encadrant_id:
        raise HTTPException(status_code=403, detail="Reserve aux encadrants")

    stagiaire = db.query(Stagiaire).filter(Stagiaire.id == stagiaire_id).first()
    if stagiaire is None:
        raise HTTPException(status_code=404, detail="Stagiaire introuvable")

    if stagiaire.encadrant_id != utilisateur_courant.encadrant_id:
        raise HTTPException(status_code=403, detail="Ce stagiaire ne vous est pas assigne")

    return stagiaire


@app.get("/moi/evaluations/a-faire")
def obtenir_evaluations_a_faire(
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "encadrant" or not utilisateur_courant.encadrant_id:
        raise HTTPException(status_code=403, detail="Reserve aux encadrants")

    stagiaires = (
        db.query(Stagiaire)
        .filter(Stagiaire.encadrant_id == utilisateur_courant.encadrant_id)
        .all()
    )

    aujourdhui = date.today()
    total_a_faire = 0
    urgentes = 0
    liste = []

    for s in stagiaires:
        a_une_evaluation_soumise = (
            db.query(Evaluation)
            .filter(Evaluation.stagiaire_id == s.id, Evaluation.statut == "soumise")
            .first()
        )
        if a_une_evaluation_soumise:
            continue

        total_a_faire += 1
        jours_restants = (s.date_fin - aujourdhui).days
        est_urgente = jours_restants <= 15
        if est_urgente:
            urgentes += 1

        liste.append({
            "stagiaire_id": s.id,
            "nom": s.nom,
            "prenom": s.prenom,
            "date_fin": s.date_fin,
            "urgente": est_urgente,
        })

    return {"total": total_a_faire, "urgentes": urgentes, "liste": liste}

@app.get("/moi/mes-stagiaires/{stagiaire_id}/presences/stats")
def obtenir_stats_presences_dun_de_mes_stagiaires(
    stagiaire_id: int,
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    verifier_stagiaire_assigne(stagiaire_id, utilisateur_courant, db)

    aujourdhui = date.today()
    debut_mois = aujourdhui.replace(day=1)

    presences_mois = (
        db.query(Presence)
        .filter(
            Presence.stagiaire_id == stagiaire_id,
            Presence.date >= debut_mois,
            Presence.date <= aujourdhui,
        )
        .all()
    )

    total_enregistrements = len(presences_mois)
    jours_presents = [p for p in presences_mois if p.present]

    taux_presence = (
        round(len(jours_presents) / total_enregistrements * 100, 1)
        if total_enregistrements > 0
        else 0
    )

    dernier = (
        db.query(Presence)
        .filter(Presence.stagiaire_id == stagiaire_id)
        .order_by(Presence.date.desc())
        .first()
    )

    return {
        "taux_presence": taux_presence,
        "total_enregistrements": total_enregistrements,
        "jours_presents": len(jours_presents),
        "derniere_presence": dernier.date if dernier else None,
    }


@app.get("/moi/mes-stagiaires/{stagiaire_id}/commentaires", response_model=list[CommentaireOut])
def obtenir_commentaires_dun_de_mes_stagiaires(
    stagiaire_id: int,
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    verifier_stagiaire_assigne(stagiaire_id, utilisateur_courant, db)

    commentaires = (
        db.query(CommentaireEncadrant)
        .filter(CommentaireEncadrant.stagiaire_id == stagiaire_id)
        .order_by(CommentaireEncadrant.date_commentaire.desc())
        .all()
    )
    return commentaires


@app.post("/moi/mes-stagiaires/{stagiaire_id}/commentaires", response_model=CommentaireOut)
def creer_commentaire_pour_un_de_mes_stagiaires(
    stagiaire_id: int,
    commentaire: CommentaireCreate,
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    verifier_stagiaire_assigne(stagiaire_id, utilisateur_courant, db)

    nouveau = CommentaireEncadrant(
        stagiaire_id=stagiaire_id,
        encadrant_id=utilisateur_courant.encadrant_id,
        titre=commentaire.titre,
        contenu=commentaire.contenu,
    )
    db.add(nouveau)
    db.commit()
    db.refresh(nouveau)
    return nouveau



    return DocumentOut.depuis_document(nouveau_document)


@app.put("/moi/mes-stagiaires/{stagiaire_id}/dates")
def modifier_dates_dun_de_mes_stagiaires(
    stagiaire_id: int,
    dates: DatesUpdate,
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    stagiaire = verifier_stagiaire_assigne(stagiaire_id, utilisateur_courant, db)

    if dates.date_fin <= dates.date_debut:
        raise HTTPException(status_code=400, detail="La date de fin doit etre apres la date de debut")

    stagiaire.date_debut = dates.date_debut
    stagiaire.date_fin = dates.date_fin
    db.commit()
    db.refresh(stagiaire)

    return {
        "id": stagiaire.id,
        "date_debut": stagiaire.date_debut,
        "date_fin": stagiaire.date_fin,
    }


@app.post("/moi/mes-stagiaires/{stagiaire_id}/incidents")
def signaler_incident_pour_un_de_mes_stagiaires(
    stagiaire_id: int,
    incident: IncidentCreate,
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    verifier_stagiaire_assigne(stagiaire_id, utilisateur_courant, db)

    notification = Notification(
        stagiaire_id=stagiaire_id,
        categorie="incident",
        urgence="haute",
        titre=incident.titre,
        contenu=incident.contenu,
    )
    db.add(notification)
    db.commit()
    db.refresh(notification)

    return {
        "id": notification.id,
        "titre": notification.titre,
        "contenu": notification.contenu,
        "date_creation": notification.date_creation,
    }


class EvaluationCreate(BaseModel):
    titre: str
    date_evaluation: date
    criteres: dict
    commentaire_global: str | None = None
    recommandations: str | None = None
    statut: str | None = "brouillon"


def calculer_note_globale(criteres: dict) -> str:
    criteres_etoiles = ["competences_techniques", "qualite_travail", "communication"]
    criteres_pourcentage = ["autonomie", "ponctualite"]

    notes_sur_20 = []
    for cle in criteres_etoiles:
        if cle in criteres and criteres[cle] is not None:
            notes_sur_20.append((criteres[cle] / 5) * 20)
    for cle in criteres_pourcentage:
        if cle in criteres and criteres[cle] is not None:
            notes_sur_20.append((criteres[cle] / 100) * 20)

    if not notes_sur_20:
        return None
    moyenne = sum(notes_sur_20) / len(notes_sur_20)
    return f"{moyenne:.1f}"


@app.get("/moi/mes-stagiaires/{stagiaire_id}/evaluations")
def obtenir_evaluations_dun_de_mes_stagiaires(
    stagiaire_id: int,
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "encadrant" or not utilisateur_courant.encadrant_id:
        raise HTTPException(status_code=403, detail="Reserve aux encadrants")

    stagiaire = db.query(Stagiaire).filter(Stagiaire.id == stagiaire_id).first()
    if stagiaire is None:
        raise HTTPException(status_code=404, detail="Stagiaire introuvable")
    if stagiaire.encadrant_id != utilisateur_courant.encadrant_id:
        raise HTTPException(status_code=403, detail="Ce stagiaire ne vous est pas assigne")

    evaluations = (
        db.query(Evaluation)
        .filter(Evaluation.stagiaire_id == stagiaire_id)
        .order_by(Evaluation.date_evaluation.desc())
        .all()
    )
    return evaluations


@app.get("/moi/mes-stagiaires/{stagiaire_id}/evaluations/{evaluation_id}")
def obtenir_une_evaluation(
    stagiaire_id: int,
    evaluation_id: int,
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "encadrant" or not utilisateur_courant.encadrant_id:
        raise HTTPException(status_code=403, detail="Reserve aux encadrants")

    stagiaire = db.query(Stagiaire).filter(Stagiaire.id == stagiaire_id).first()
    if stagiaire is None:
        raise HTTPException(status_code=404, detail="Stagiaire introuvable")
    if stagiaire.encadrant_id != utilisateur_courant.encadrant_id:
        raise HTTPException(status_code=403, detail="Ce stagiaire ne vous est pas assigne")

    evaluation = (
        db.query(Evaluation)
        .filter(Evaluation.id == evaluation_id, Evaluation.stagiaire_id == stagiaire_id)
        .first()
    )
    if evaluation is None:
        raise HTTPException(status_code=404, detail="Evaluation introuvable")
    return evaluation


@app.post("/moi/mes-stagiaires/{stagiaire_id}/evaluations")
def creer_evaluation(
    stagiaire_id: int,
    donnees: EvaluationCreate,
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "encadrant" or not utilisateur_courant.encadrant_id:
        raise HTTPException(status_code=403, detail="Reserve aux encadrants")

    stagiaire = db.query(Stagiaire).filter(Stagiaire.id == stagiaire_id).first()
    if stagiaire is None:
        raise HTTPException(status_code=404, detail="Stagiaire introuvable")
    if stagiaire.encadrant_id != utilisateur_courant.encadrant_id:
        raise HTTPException(status_code=403, detail="Ce stagiaire ne vous est pas assigne")

    note_globale = calculer_note_globale(donnees.criteres)

    nouvelle_evaluation = Evaluation(
        stagiaire_id=stagiaire_id,
        encadrant_id=utilisateur_courant.encadrant_id,
        titre=donnees.titre,
        note=note_globale,
        date_evaluation=donnees.date_evaluation,
        criteres=donnees.criteres,
        commentaire_global=donnees.commentaire_global,
        recommandations=donnees.recommandations,
        statut=donnees.statut,
    )
    db.add(nouvelle_evaluation)
    db.commit()
    db.refresh(nouvelle_evaluation)
    return nouvelle_evaluation


@app.put("/moi/mes-stagiaires/{stagiaire_id}/evaluations/{evaluation_id}")
def modifier_evaluation(
    stagiaire_id: int,
    evaluation_id: int,
    donnees: EvaluationCreate,
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "encadrant" or not utilisateur_courant.encadrant_id:
        raise HTTPException(status_code=403, detail="Reserve aux encadrants")

    evaluation = (
        db.query(Evaluation)
        .filter(Evaluation.id == evaluation_id, Evaluation.stagiaire_id == stagiaire_id)
        .first()
    )
    if evaluation is None:
        raise HTTPException(status_code=404, detail="Evaluation introuvable")
    if evaluation.encadrant_id != utilisateur_courant.encadrant_id:
        raise HTTPException(status_code=403, detail="Cette evaluation ne vous appartient pas")

    note_globale = calculer_note_globale(donnees.criteres)

    evaluation.titre = donnees.titre
    evaluation.date_evaluation = donnees.date_evaluation
    evaluation.criteres = donnees.criteres
    evaluation.commentaire_global = donnees.commentaire_global
    evaluation.recommandations = donnees.recommandations
    evaluation.statut = donnees.statut
    evaluation.note = note_globale

    db.commit()
    db.refresh(evaluation)
    return evaluation


@app.get("/moi/encadrant")
def obtenir_mon_encadrant(
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "stagiaire" or not utilisateur_courant.stagiaire_id:
        raise HTTPException(status_code=403, detail="Reserve aux stagiaires")

    stagiaire = db.query(Stagiaire).filter(Stagiaire.id == utilisateur_courant.stagiaire_id).first()
    if stagiaire is None or stagiaire.encadrant_id is None:
        raise HTTPException(status_code=404, detail="Aucun encadrant assigne")

    encadrant = db.query(Encadrant).filter(Encadrant.id == stagiaire.encadrant_id).first()
    if encadrant is None:
        raise HTTPException(status_code=404, detail="Encadrant introuvable")

    departement = None
    if encadrant.departement_id:
        departement = db.query(Departement).filter(Departement.id == encadrant.departement_id).first()

    nb_stagiaires_encadres = db.query(Stagiaire).filter(Stagiaire.encadrant_id == encadrant.id).count()

    prochaine_reunion = (
        db.query(Reunion)
        .filter(
            Reunion.stagiaire_id == stagiaire.id,
            Reunion.encadrant_id == encadrant.id,
            Reunion.statut == "a_venir",
        )
        .order_by(Reunion.date_reunion.asc(), Reunion.heure.asc())
        .first()
    )

    prochain_rendez_vous = None
    if prochaine_reunion is not None:
        prochain_rendez_vous = {
            "date": prochaine_reunion.date_reunion.isoformat() if prochaine_reunion.date_reunion else None,
            "heure": prochaine_reunion.heure,
            "objet": prochaine_reunion.objet,
            "type_reunion": prochaine_reunion.type_reunion,
        }

    return {
        "id": encadrant.id,
        "nom": encadrant.nom,
        "prenom": encadrant.prenom,
        "email": encadrant.email,
        "fonction": encadrant.fonction,
        "telephone": encadrant.telephone,
        "bureau": encadrant.bureau,
        "horaires_disponibilite": encadrant.horaires_disponibilite,
        "photo_url": encadrant.photo_url,
        "departement": departement.nom if departement else None,
        "nb_stagiaires_encadres": nb_stagiaires_encadres,
        "prochain_rendez_vous": prochain_rendez_vous,
    }


@app.get("/moi/encadrant/commentaires")
def obtenir_commentaires_encadrant(
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "stagiaire" or not utilisateur_courant.stagiaire_id:
        raise HTTPException(status_code=403, detail="Reserve aux stagiaires")

    commentaires = (
        db.query(CommentaireEncadrant)
        .filter(CommentaireEncadrant.stagiaire_id == utilisateur_courant.stagiaire_id)
        .order_by(CommentaireEncadrant.date_commentaire.desc())
        .all()
    )
    return commentaires


@app.get("/moi/encadrant/evaluations")
def obtenir_evaluations(
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "stagiaire" or not utilisateur_courant.stagiaire_id:
        raise HTTPException(status_code=403, detail="Reserve aux stagiaires")

    evaluations = (
        db.query(Evaluation)
        .filter(Evaluation.stagiaire_id == utilisateur_courant.stagiaire_id)
        .order_by(Evaluation.date_evaluation.desc())
        .all()
    )
    return evaluations


class MessageCreate(BaseModel):
    contenu: str
    type_message: str = "message"


class MessageEncadrantCreate(BaseModel):
    stagiaire_id: int
    contenu: str
    type_message: str = "message"


class ReunionCreate(BaseModel):
    stagiaire_id: int
    date_reunion: date
    heure: str
    type_reunion: str = "presentiel"
    lieu_ou_lien: str | None = None
    objet: str
    notes: str | None = None


class ReunionUpdate(BaseModel):
    date_reunion: date | None = None
    heure: str | None = None
    type_reunion: str | None = None
    lieu_ou_lien: str | None = None
    objet: str | None = None
    notes: str | None = None


def calculer_statut_reunion(reunion, maintenant=None):
    if reunion.statut == "annulee":
        return "annulee"
    maintenant = maintenant or datetime.now()
    try:
        heure_debut = datetime.strptime(reunion.heure, "%H:%M").time()
    except (ValueError, TypeError):
        heure_debut = datetime.min.time()
    debut = datetime.combine(reunion.date_reunion, heure_debut)
    fin = debut + timedelta(hours=1)
    if maintenant < debut:
        return "a_venir"
    if debut <= maintenant <= fin:
        return "en_cours"
    return "terminee"


@app.get("/moi/messages")
def obtenir_mes_messages(
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "stagiaire" or not utilisateur_courant.stagiaire_id:
        raise HTTPException(status_code=403, detail="Reserve aux stagiaires")

    messages = (
        db.query(Message)
        .filter(Message.stagiaire_id == utilisateur_courant.stagiaire_id)
        .order_by(Message.date_envoi.asc())
        .all()
    )
    return messages


@app.post("/moi/messages")
def envoyer_message(
    donnees: MessageCreate,
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "stagiaire" or not utilisateur_courant.stagiaire_id:
        raise HTTPException(status_code=403, detail="Reserve aux stagiaires")

    stagiaire = db.query(Stagiaire).filter(Stagiaire.id == utilisateur_courant.stagiaire_id).first()
    if stagiaire is None or stagiaire.encadrant_id is None:
        raise HTTPException(status_code=404, detail="Aucun encadrant assigne")

    nouveau_message = Message(
        stagiaire_id=utilisateur_courant.stagiaire_id,
        encadrant_id=stagiaire.encadrant_id,
        expediteur="stagiaire",
        type_message=donnees.type_message,
        contenu=donnees.contenu,
        lu=False,
    )
    db.add(nouveau_message)
    db.commit()
    db.refresh(nouveau_message)
    return nouveau_message


@app.get("/encadrant/messages")
def lister_conversations_encadrant(
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "encadrant" or not utilisateur_courant.encadrant_id:
        raise HTTPException(status_code=403, detail="Reserve aux encadrants")

    stagiaires = (
        db.query(Stagiaire)
        .filter(Stagiaire.encadrant_id == utilisateur_courant.encadrant_id)
        .all()
    )

    conversations = []
    for s in stagiaires:
        dernier_message = (
            db.query(Message)
            .filter(Message.stagiaire_id == s.id)
            .order_by(Message.date_envoi.desc())
            .first()
        )
        conversations.append({
            "stagiaire_id": s.id,
            "prenom": s.prenom,
            "nom": s.nom,
            "dernier_message": dernier_message.contenu if dernier_message else None,
            "date_dernier_message": dernier_message.date_envoi.isoformat() if dernier_message else None,
            "non_lu": (
                db.query(Message)
                .filter(
                    Message.stagiaire_id == s.id,
                    Message.expediteur == "stagiaire",
                    Message.lu == False,
                )
                .count()
                > 0
            ),
        })

    conversations.sort(
        key=lambda c: c["date_dernier_message"] or "",
        reverse=True,
    )
    return conversations


@app.get("/encadrant/messages/non-lus")
def compter_messages_non_lus_encadrant(
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "encadrant" or not utilisateur_courant.encadrant_id:
        raise HTTPException(status_code=403, detail="Reserve aux encadrants")

    stagiaires_ids = [
        s.id
        for s in db.query(Stagiaire.id)
        .filter(Stagiaire.encadrant_id == utilisateur_courant.encadrant_id)
        .all()
    ]

    count = (
        db.query(Message)
        .filter(
            Message.stagiaire_id.in_(stagiaires_ids),
            Message.expediteur == "stagiaire",
            Message.lu == False,
        )
        .count()
    )
    return {"non_lus": count}


@app.get("/encadrant/messages/{stagiaire_id}")
def lister_messages_stagiaire_encadrant(
    stagiaire_id: int,
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "encadrant" or not utilisateur_courant.encadrant_id:
        raise HTTPException(status_code=403, detail="Reserve aux encadrants")

    stagiaire = db.query(Stagiaire).filter(
        Stagiaire.id == stagiaire_id,
        Stagiaire.encadrant_id == utilisateur_courant.encadrant_id,
    ).first()
    if stagiaire is None:
        raise HTTPException(status_code=404, detail="Stagiaire introuvable")

    messages = (
        db.query(Message)
        .filter(Message.stagiaire_id == stagiaire_id)
        .order_by(Message.date_envoi.asc())
        .all()
    )

    messages_serialises = [
        {
            "id": m.id,
            "stagiaire_id": m.stagiaire_id,
            "encadrant_id": m.encadrant_id,
            "expediteur": m.expediteur,
            "type_message": m.type_message,
            "contenu": m.contenu,
            "date_envoi": m.date_envoi.isoformat() if m.date_envoi else None,
            "lu": m.lu,
            "piece_jointe_nom": m.piece_jointe_nom,
            "piece_jointe_url": m.piece_jointe_url,
            "piece_jointe_taille": m.piece_jointe_taille,
        }
        for m in messages
    ]

    db.query(Message).filter(
        Message.stagiaire_id == stagiaire_id,
        Message.expediteur == "stagiaire",
        Message.lu == False,
    ).update({"lu": True})
    db.commit()

    return messages_serialises


@app.post("/encadrant/messages")
def envoyer_message_encadrant(
    donnees: MessageEncadrantCreate,
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "encadrant" or not utilisateur_courant.encadrant_id:
        raise HTTPException(status_code=403, detail="Reserve aux encadrants")

    stagiaire = db.query(Stagiaire).filter(
        Stagiaire.id == donnees.stagiaire_id,
        Stagiaire.encadrant_id == utilisateur_courant.encadrant_id,
    ).first()
    if stagiaire is None:
        raise HTTPException(status_code=404, detail="Stagiaire introuvable")

    nouveau_message = Message(
        stagiaire_id=donnees.stagiaire_id,
        encadrant_id=utilisateur_courant.encadrant_id,
        expediteur="encadrant",
        type_message=donnees.type_message,
        contenu=donnees.contenu,
        lu=False,
    )
    db.add(nouveau_message)
    db.commit()
    db.refresh(nouveau_message)
    return nouveau_message


@app.post("/moi/messages/piece-jointe")
async def envoyer_message_avec_piece_jointe_stagiaire(
    contenu: str = Form(""),
    fichier: UploadFile = File(...),
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "stagiaire" or not utilisateur_courant.stagiaire_id:
        raise HTTPException(status_code=403, detail="Reserve aux stagiaires")

    stagiaire = db.query(Stagiaire).filter(Stagiaire.id == utilisateur_courant.stagiaire_id).first()
    if stagiaire is None or stagiaire.encadrant_id is None:
        raise HTTPException(status_code=404, detail="Aucun encadrant assigne")

    extension = os.path.splitext(fichier.filename)[1].lower()
    if extension not in EXTENSIONS_AUTORISEES:
        raise HTTPException(status_code=400, detail="Type de fichier non autorise")

    contenu_fichier = await fichier.read()
    if len(contenu_fichier) > TAILLE_MAX_OCTETS:
        raise HTTPException(status_code=400, detail="Fichier trop volumineux (max 10 Mo)")

    nom_unique = f"{uuid.uuid4().hex}{extension}"
    chemin_disque = os.path.join(UPLOAD_DIR, nom_unique)
    with open(chemin_disque, "wb") as f:
        f.write(contenu_fichier)

    nouveau_message_pj = Message(
        stagiaire_id=utilisateur_courant.stagiaire_id,
        encadrant_id=stagiaire.encadrant_id,
        expediteur="stagiaire",
        type_message="fichier",
        contenu=contenu.strip() if contenu.strip() else fichier.filename,
        lu=False,
        piece_jointe_nom=fichier.filename,
        piece_jointe_url=f"/uploads/{nom_unique}",
        piece_jointe_taille=len(contenu_fichier),
    )
    db.add(nouveau_message_pj)
    db.commit()
    db.refresh(nouveau_message_pj)
    return nouveau_message_pj


@app.post("/encadrant/messages/piece-jointe")
async def envoyer_message_avec_piece_jointe_encadrant(
    stagiaire_id: int = Form(...),
    contenu: str = Form(""),
    fichier: UploadFile = File(...),
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "encadrant" or not utilisateur_courant.encadrant_id:
        raise HTTPException(status_code=403, detail="Reserve aux encadrants")

    stagiaire = db.query(Stagiaire).filter(
        Stagiaire.id == stagiaire_id,
        Stagiaire.encadrant_id == utilisateur_courant.encadrant_id,
    ).first()
    if stagiaire is None:
        raise HTTPException(status_code=404, detail="Stagiaire introuvable")

    extension = os.path.splitext(fichier.filename)[1].lower()
    if extension not in EXTENSIONS_AUTORISEES:
        raise HTTPException(status_code=400, detail="Type de fichier non autorise")

    contenu_fichier = await fichier.read()
    if len(contenu_fichier) > TAILLE_MAX_OCTETS:
        raise HTTPException(status_code=400, detail="Fichier trop volumineux (max 10 Mo)")

    nom_unique = f"{uuid.uuid4().hex}{extension}"
    chemin_disque = os.path.join(UPLOAD_DIR, nom_unique)
    with open(chemin_disque, "wb") as f:
        f.write(contenu_fichier)

    nouveau_message_pj = Message(
        stagiaire_id=stagiaire_id,
        encadrant_id=utilisateur_courant.encadrant_id,
        expediteur="encadrant",
        type_message="fichier",
        contenu=contenu.strip() if contenu.strip() else fichier.filename,
        lu=False,
        piece_jointe_nom=fichier.filename,
        piece_jointe_url=f"/uploads/{nom_unique}",
        piece_jointe_taille=len(contenu_fichier),
    )
    db.add(nouveau_message_pj)
    db.commit()
    db.refresh(nouveau_message_pj)
    return nouveau_message_pj


@app.post("/encadrant/reunions")
def creer_reunion(
    donnees: ReunionCreate,
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "encadrant" or not utilisateur_courant.encadrant_id:
        raise HTTPException(status_code=403, detail="Reserve aux encadrants")

    stagiaire = db.query(Stagiaire).filter(
        Stagiaire.id == donnees.stagiaire_id,
        Stagiaire.encadrant_id == utilisateur_courant.encadrant_id,
    ).first()
    if stagiaire is None:
        raise HTTPException(status_code=404, detail="Stagiaire introuvable")

    nouvelle_reunion = Reunion(
        stagiaire_id=donnees.stagiaire_id,
        encadrant_id=utilisateur_courant.encadrant_id,
        date_reunion=donnees.date_reunion,
        heure=donnees.heure,
        type_reunion=donnees.type_reunion,
        lieu_ou_lien=donnees.lieu_ou_lien,
        objet=donnees.objet,
        notes=donnees.notes,
    )
    db.add(nouvelle_reunion)

    date_fr = donnees.date_reunion.strftime("%d/%m/%Y")
    nouvelle_notification = Notification(
        stagiaire_id=donnees.stagiaire_id,
        categorie="reunion",
        urgence="normale",
        titre="Nouvelle réunion planifiée",
        contenu=f"{donnees.objet} — le {date_fr} à {donnees.heure}",
        lu=False,
    )
    db.add(nouvelle_notification)

    db.commit()
    db.refresh(nouvelle_reunion)
    return nouvelle_reunion


@app.put("/encadrant/reunions/{reunion_id}")
def modifier_reunion(
    reunion_id: int,
    donnees: ReunionUpdate,
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "encadrant" or not utilisateur_courant.encadrant_id:
        raise HTTPException(status_code=403, detail="Reserve aux encadrants")

    reunion = db.query(Reunion).filter(
        Reunion.id == reunion_id,
        Reunion.encadrant_id == utilisateur_courant.encadrant_id,
    ).first()
    if reunion is None:
        raise HTTPException(status_code=404, detail="Reunion introuvable")
    if reunion.statut == "annulee":
        raise HTTPException(status_code=400, detail="Impossible de modifier une reunion annulee")

    for champ, valeur in donnees.dict(exclude_unset=True).items():
        setattr(reunion, champ, valeur)

    db.commit()
    db.refresh(reunion)
    return reunion


@app.put("/encadrant/reunions/{reunion_id}/annuler")
def annuler_reunion(
    reunion_id: int,
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "encadrant" or not utilisateur_courant.encadrant_id:
        raise HTTPException(status_code=403, detail="Reserve aux encadrants")

    reunion = db.query(Reunion).filter(
        Reunion.id == reunion_id,
        Reunion.encadrant_id == utilisateur_courant.encadrant_id,
    ).first()
    if reunion is None:
        raise HTTPException(status_code=404, detail="Reunion introuvable")

    reunion.statut = "annulee"

    nouvelle_notification = Notification(
        stagiaire_id=reunion.stagiaire_id,
        categorie="reunion",
        urgence="normale",
        titre="Reunion annulee",
        contenu=f"{reunion.objet} prevue le {reunion.date_reunion.strftime('%d/%m/%Y')} a ete annulee.",
        lu=False,
    )
    db.add(nouvelle_notification)

    db.commit()
    db.refresh(reunion)
    return {"ok": True, "statut": reunion.statut}


@app.put("/encadrant/reunions/{reunion_id}/rappel")
def envoyer_rappel_reunion(
    reunion_id: int,
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "encadrant" or not utilisateur_courant.encadrant_id:
        raise HTTPException(status_code=403, detail="Reserve aux encadrants")

    reunion = db.query(Reunion).filter(
        Reunion.id == reunion_id,
        Reunion.encadrant_id == utilisateur_courant.encadrant_id,
    ).first()
    if reunion is None:
        raise HTTPException(status_code=404, detail="Reunion introuvable")

    if reunion.statut != "a_venir":
        raise HTTPException(status_code=400, detail="Le rappel n'est disponible que pour les reunions a venir")

    date_fr = reunion.date_reunion.strftime("%d/%m/%Y")

    nouvelle_notification = Notification(
        stagiaire_id=reunion.stagiaire_id,
        categorie="reunion",
        urgence="normale",
        titre="Rappel de reunion",
        contenu=f"Rappel : {reunion.objet} prevue le {date_fr} a {reunion.heure}.",
        lu=False,
    )
    db.add(nouvelle_notification)
    db.commit()

    return {"ok": True, "notification_envoyee": True}


@app.get("/moi/prochaine-reunion")
def obtenir_prochaine_reunion(
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "stagiaire" or not utilisateur_courant.stagiaire_id:
        raise HTTPException(status_code=403, detail="Reserve aux stagiaires")

    aujourdhui = date.today()
    reunion = (
        db.query(Reunion)
        .filter(
            Reunion.stagiaire_id == utilisateur_courant.stagiaire_id,
            Reunion.date_reunion >= aujourdhui,
        )
        .order_by(Reunion.date_reunion.asc(), Reunion.heure.asc())
        .first()
    )
    if reunion is None:
        return None

    encadrant = db.query(Encadrant).filter(Encadrant.id == reunion.encadrant_id).first()

    return {
        "titre": reunion.objet,
        "date": reunion.date_reunion.isoformat(),
        "heure": reunion.heure,
        "avec": f"{encadrant.prenom} {encadrant.nom}" if encadrant else None,
    }


@app.get("/encadrant/reunions")
def lister_reunions_encadrant(
    stagiaire_id: int | None = None,
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "encadrant" or not utilisateur_courant.encadrant_id:
        raise HTTPException(status_code=403, detail="Reserve aux encadrants")

    requete = db.query(Reunion).filter(Reunion.encadrant_id == utilisateur_courant.encadrant_id)
    if stagiaire_id is not None:
        requete = requete.filter(Reunion.stagiaire_id == stagiaire_id)

    reunions = requete.order_by(Reunion.date_reunion.asc(), Reunion.heure.asc()).all()

    maintenant = datetime.now()
    resultats = []
    for r in reunions:
        stagiaire = db.query(Stagiaire).filter(Stagiaire.id == r.stagiaire_id).first()
        resultats.append({
            "id": r.id,
            "stagiaire_id": r.stagiaire_id,
            "stagiaire_nom": stagiaire.nom if stagiaire else None,
            "stagiaire_prenom": stagiaire.prenom if stagiaire else None,
            "date_reunion": r.date_reunion.isoformat(),
            "heure": r.heure,
            "type_reunion": r.type_reunion,
            "lieu_ou_lien": r.lieu_ou_lien,
            "objet": r.objet,
            "notes": r.notes,
            "statut": calculer_statut_reunion(r, maintenant),
        })
    return resultats


@app.get("/moi/messages/non-lus")
def compter_messages_non_lus(
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "stagiaire" or not utilisateur_courant.stagiaire_id:
        raise HTTPException(status_code=403, detail="Reserve aux stagiaires")

    count = (
        db.query(Message)
        .filter(
            Message.stagiaire_id == utilisateur_courant.stagiaire_id,
            Message.expediteur == "encadrant",
            Message.lu == False,
        )
        .count()
    )
    return {"non_lus": count}


@app.put("/moi/messages/marquer-lus")
def marquer_mes_messages_lus(
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "stagiaire" or not utilisateur_courant.stagiaire_id:
        raise HTTPException(status_code=403, detail="Reserve aux stagiaires")

    db.query(Message).filter(
        Message.stagiaire_id == utilisateur_courant.stagiaire_id,
        Message.expediteur == "encadrant",
        Message.lu == False,
    ).update({"lu": True})
    db.commit()
    return {"ok": True}


@app.get("/moi/notifications")
def obtenir_mes_notifications(
    categorie: str | None = None,
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "stagiaire" or not utilisateur_courant.stagiaire_id:
        raise HTTPException(status_code=403, detail="Reserve aux stagiaires")

    requete = db.query(Notification).filter(
        Notification.stagiaire_id == utilisateur_courant.stagiaire_id
    )

    if categorie and categorie != "tous":
        requete = requete.filter(Notification.categorie == categorie)

    notifications = requete.order_by(Notification.date_creation.desc()).all()
    return notifications


@app.get("/moi/notifications/non-lues")
def compter_notifications_non_lues(
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "stagiaire" or not utilisateur_courant.stagiaire_id:
        raise HTTPException(status_code=403, detail="Reserve aux stagiaires")

    count = (
        db.query(Notification)
        .filter(
            Notification.stagiaire_id == utilisateur_courant.stagiaire_id,
            Notification.lu == False,
        )
        .count()
    )
    return {"non_lues": count}


@app.put("/moi/notifications/{notification_id}/lu")
def marquer_notification_lue(
    notification_id: int,
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "stagiaire" or not utilisateur_courant.stagiaire_id:
        raise HTTPException(status_code=403, detail="Reserve aux stagiaires")

    notification = (
        db.query(Notification)
        .filter(
            Notification.id == notification_id,
            Notification.stagiaire_id == utilisateur_courant.stagiaire_id,
        )
        .first()
    )
    if not notification:
        raise HTTPException(status_code=404, detail="Notification introuvable")

    notification.lu = True
    db.commit()


@app.get("/encadrant/notifications")
def obtenir_notifications_encadrant(
    categorie: str | None = None,
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "encadrant" or not utilisateur_courant.encadrant_id:
        raise HTTPException(status_code=403, detail="Reserve aux encadrants")

    requete = db.query(Notification).filter(
        Notification.encadrant_id == utilisateur_courant.encadrant_id
    )

    if categorie and categorie != "tous":
        requete = requete.filter(Notification.categorie == categorie)

    notifications = requete.order_by(Notification.date_creation.desc()).all()
    return notifications


@app.get("/encadrant/notifications/non-lues")
def compter_notifications_non_lues_encadrant(
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "encadrant" or not utilisateur_courant.encadrant_id:
        raise HTTPException(status_code=403, detail="Reserve aux encadrants")

    count = (
        db.query(Notification)
        .filter(
            Notification.encadrant_id == utilisateur_courant.encadrant_id,
            Notification.lu == False,
        )
        .count()
    )
    return {"non_lues": count}


@app.put("/encadrant/notifications/{notification_id}/lu")
def marquer_notification_lue_encadrant(
    notification_id: int,
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "encadrant" or not utilisateur_courant.encadrant_id:
        raise HTTPException(status_code=403, detail="Reserve aux encadrants")

    notification = (
        db.query(Notification)
        .filter(
            Notification.id == notification_id,
            Notification.encadrant_id == utilisateur_courant.encadrant_id,
        )
        .first()
    )
    if not notification:
        raise HTTPException(status_code=404, detail="Notification introuvable")

    notification.lu = True
    db.commit()
    return {"ok": True}


@app.put("/encadrant/notifications/tout-lire")
def marquer_toutes_notifications_lues_encadrant(
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "encadrant" or not utilisateur_courant.encadrant_id:
        raise HTTPException(status_code=403, detail="Reserve aux encadrants")

    db.query(Notification).filter(
        Notification.encadrant_id == utilisateur_courant.encadrant_id,
        Notification.lu == False,
    ).update({"lu": True})
    db.commit()
    return {"ok": True}


@app.delete("/encadrant/notifications")
def effacer_toutes_notifications_encadrant(
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "encadrant" or not utilisateur_courant.encadrant_id:
        raise HTTPException(status_code=403, detail="Reserve aux encadrants")

    db.query(Notification).filter(
        Notification.encadrant_id == utilisateur_courant.encadrant_id
    ).delete()
    db.commit()
    return {"ok": True}


def generer_rappels_evaluation(db: Session):
    """
    Parcourt les stagiaires en cours et cree une notification de rappel
    pour l'encadrant si le stage a depasse la mi-parcours et qu'aucune
    evaluation n'existe encore pour ce stagiaire.
    """
    aujourdhui = date.today()

    stagiaires_en_cours = db.query(Stagiaire).filter(Stagiaire.statut == "en_cours").all()

    for stagiaire in stagiaires_en_cours:
        if not stagiaire.encadrant_id:
            continue

        duree_totale = (stagiaire.date_fin - stagiaire.date_debut).days
        if duree_totale <= 0:
            continue

        date_mi_parcours = stagiaire.date_debut + timedelta(days=duree_totale // 2)

        if aujourdhui < date_mi_parcours or aujourdhui > stagiaire.date_fin:
            continue

        evaluation_existante = (
            db.query(Evaluation)
            .filter(Evaluation.stagiaire_id == stagiaire.id)
            .first()
        )
        if evaluation_existante:
            continue

        rappel_recent = (
            db.query(Notification)
            .filter(
                Notification.stagiaire_id == stagiaire.id,
                Notification.categorie == "evaluation",
                Notification.date_creation >= datetime.utcnow() - timedelta(days=3),
            )
            .first()
        )
        if rappel_recent:
            continue

        nouvelle_notification = Notification(
            encadrant_id=stagiaire.encadrant_id,
            stagiaire_id=stagiaire.id,
            categorie="evaluation",
            urgence="haute",
            titre="Rappel d'évaluation",
            contenu=(
                f"L'évaluation de mi-parcours de {stagiaire.prenom} {stagiaire.nom} "
                f"doit être validée avant le {stagiaire.date_fin.strftime('%d/%m/%Y')}."
            ),
        )
        db.add(nouvelle_notification)

    db.commit()


@app.post("/encadrant/taches/verifier-rappels")
def verifier_rappels_evaluation_maintenant(
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "encadrant":
        raise HTTPException(status_code=403, detail="Reserve aux encadrants")

    generer_rappels_evaluation(db)
    return {"ok": True}


def generer_validations_presence(db: Session):
    """
    Pour chaque encadrant, verifie si les presences de la semaine derniere
    (lundi-vendredi) sont completes pour tous ses stagiaires actifs.
    Si oui, cree une notification de confirmation.
    """
    aujourdhui = date.today()

    lundi_cette_semaine = aujourdhui - timedelta(days=aujourdhui.weekday())
    lundi_semaine_derniere = lundi_cette_semaine - timedelta(days=7)
    jours_semaine_derniere = [lundi_semaine_derniere + timedelta(days=i) for i in range(5)]

    encadrants_ids = [row[0] for row in db.query(Encadrant.id).all()]

    for encadrant_id in encadrants_ids:
        stagiaires_actifs = (
            db.query(Stagiaire)
            .filter(Stagiaire.encadrant_id == encadrant_id, Stagiaire.statut == "en_cours")
            .all()
        )
        if not stagiaires_actifs:
            continue

        tout_est_complet = True
        for stagiaire in stagiaires_actifs:
            nb_presences = (
                db.query(Presence)
                .filter(
                    Presence.stagiaire_id == stagiaire.id,
                    Presence.date.in_(jours_semaine_derniere),
                )
                .count()
            )
            if nb_presences < len(jours_semaine_derniere):
                tout_est_complet = False
                break

        if not tout_est_complet:
            continue

        notification_existante = (
            db.query(Notification)
            .filter(
                Notification.encadrant_id == encadrant_id,
                Notification.categorie == "presence",
                Notification.date_creation >= datetime.utcnow() - timedelta(days=7),
            )
            .first()
        )
        if notification_existante:
            continue

        nouvelle_notification = Notification(
            encadrant_id=encadrant_id,
            categorie="presence",
            urgence="normale",
            titre="Validation de présence",
            contenu="Les feuilles de présence de la semaine dernière ont été validées automatiquement.",
        )
        db.add(nouvelle_notification)

    db.commit()


@app.post("/encadrant/taches/verifier-presences")
def verifier_validations_presence_maintenant(
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "encadrant":
        raise HTTPException(status_code=403, detail="Reserve aux encadrants")

    generer_validations_presence(db)
    return {"ok": True}


@app.on_event("startup")
def tache_demarrage_rappels():
    db = SessionLocal()
    try:
        generer_rappels_evaluation(db)
        generer_validations_presence(db)
    finally:
        db.close()


@app.put("/moi/notifications/tout-lire")
def marquer_toutes_notifications_lues(
    utilisateur_courant: Utilisateur = Depends(obtenir_utilisateur_courant),
    db: Session = Depends(get_db),
):
    if utilisateur_courant.role != "stagiaire" or not utilisateur_courant.stagiaire_id:
        raise HTTPException(status_code=403, detail="Reserve aux stagiaires")

    (
        db.query(Notification)
        .filter(
            Notification.stagiaire_id == utilisateur_courant.stagiaire_id,
            Notification.lu == False,
        )
        .update({"lu": True})
    )
    db.commit()
    return {"message": "Toutes les notifications ont ete marquees comme lues"}


# ============================================================
# Demandes de stage (candidatures externes)
# ============================================================

def _log_activite(db: Session, stagiaire_id: int, action: str, statut: str | None = None):
    """
    Enregistre une activité liée à un stagiaire. Ne fait PAS de commit :
    c'est la route appelante qui commit une seule fois (objet principal +
    activité ensemble) pour éviter d'expirer l'objet déjà chargé.
    """
    nouvelle_activite = Activite(
        stagiaire_id=stagiaire_id,
        action=action,
        statut=statut,
    )
    db.add(nouvelle_activite)


EXTENSIONS_AUTORISEES_CANDIDATURE = {".pdf", ".doc", ".docx"}


@app.post("/demandes-stage/upload-document")
async def televerser_document_candidature(fichier: UploadFile = File(...)):
    """
    Route PUBLIQUE : upload du CV ou de la lettre de motivation par un
    candidat externe, avant la création de la DemandeStage elle-même.
    Retourne l'URL à stocker dans cv_url / lettre_motivation_url.
    """
    extension = os.path.splitext(fichier.filename or "")[1].lower()
    if extension not in EXTENSIONS_AUTORISEES_CANDIDATURE:
        raise HTTPException(
            status_code=400,
            detail=f"Extension non autorisée. Formats acceptés : {', '.join(EXTENSIONS_AUTORISEES_CANDIDATURE)}",
        )
    nom_unique = f"{uuid.uuid4().hex}{extension}"
    chemin_disque = os.path.join(UPLOAD_DIR, nom_unique)
    with open(chemin_disque, "wb") as buffer:
        shutil.copyfileobj(fichier.file, buffer)
    return {"url": f"/uploads/{nom_unique}"}


@app.get("/demandes-stage")
def liste_demandes_stage(
    db: Session = Depends(get_db),
    utilisateur_courant: Utilisateur = Depends(exiger_role("rh", "admin_rh")),
):
    return db.query(DemandeStage).order_by(DemandeStage.id.desc()).all()


# --- Suivi PUBLIC d'une candidature (portail candidat, sans authentification) ---
# IMPORTANT : cette route doit être déclarée AVANT "/demandes-stage/{demande_id}"
# pour que FastAPI ne confonde pas "suivi" avec un id.
@app.get("/demandes-stage/suivi")
def suivi_demande_stage(id: int, email: str, db: Session = Depends(get_db)):
    """
    Route PUBLIQUE utilisée par la page "Suivre ma candidature". Le candidat
    doit fournir à la fois l'id de son dossier ET l'email utilisé lors de sa
    candidature, pour éviter qu'un id deviné donne accès au dossier de
    quelqu'un d'autre.
    """
    demande = (
        db.query(DemandeStage)
        .filter(DemandeStage.id == id, DemandeStage.email.ilike(email.strip()))
        .first()
    )
    if demande is None:
        raise HTTPException(
            status_code=404,
            detail="Aucune candidature trouvée avec ces informations",
        )
    return demande


@app.get("/demandes-stage/{demande_id}")
def detail_demande_stage(
    demande_id: int,
    db: Session = Depends(get_db),
    utilisateur_courant: Utilisateur = Depends(exiger_role("admin_rh")),
):
    demande = db.query(DemandeStage).filter(DemandeStage.id == demande_id).first()
    if demande is None:
        raise HTTPException(status_code=404, detail="Demande de stage non trouvée")
    return demande


@app.post("/demandes-stage")
def creer_demande_stage(demande: DemandeStageCreate, db: Session = Depends(get_db)):
    """
    Route PUBLIQUE (pas d'authentification) : un candidat externe soumet sa
    candidature. Le statut est toujours "en_attente" au départ.
    """
    nouvelle = DemandeStage(
        **demande.dict(),
        statut="en_attente",
        date_creation=datetime.now().isoformat(),
    )
    db.add(nouvelle)
    db.commit()
    db.refresh(nouvelle)
    return nouvelle


@app.put("/demandes-stage/{demande_id}")
def modifier_demande_stage(
    demande_id: int,
    demande_maj: DemandeStageUpdate,
    db: Session = Depends(get_db),
    utilisateur_courant: Utilisateur = Depends(exiger_role("admin_rh")),
):
    """
    Route RH : fait avancer une candidature dans son workflow. Seuls les
    champs envoyés (non None) sont mis à jour.
    """
    demande = db.query(DemandeStage).filter(DemandeStage.id == demande_id).first()
    if demande is None:
        raise HTTPException(status_code=404, detail="Demande de stage non trouvée")

    for champ, valeur in demande_maj.dict(exclude_unset=True).items():
        setattr(demande, champ, valeur)

    db.commit()
    db.refresh(demande)
    return demande


@app.delete("/demandes-stage/{demande_id}")
def supprimer_demande_stage(
    demande_id: int,
    db: Session = Depends(get_db),
    utilisateur_courant: Utilisateur = Depends(exiger_role("admin_rh")),
):
    obj = db.query(DemandeStage).filter(DemandeStage.id == demande_id).first()
    if obj is None:
        raise HTTPException(status_code=404, detail="Demande de stage non trouvée")
    db.delete(obj)
    db.commit()
    return {"message": f"Demande de stage {demande_id} supprimée"}


@app.post("/demandes-stage/{demande_id}/convertir")
def convertir_demande_en_stagiaire(
    demande_id: int,
    db: Session = Depends(get_db),
    utilisateur_courant: Utilisateur = Depends(exiger_role("admin_rh")),
):
    """
    Convertit une candidature en fiche Stagiaire. La demande d'origine est
    CONSERVÉE (pas supprimée) pour garder un historique, et on note l'id du
    stagiaire créé sur stagiaire_id_cree pour la traçabilité.

    IMPORTANT : chez nous, Stagiaire utilise etablissement_id / departement_id
    (ForeignKey), alors que DemandeStage stocke etablissements / departements
    en texte libre. On cherche donc l'Etablissement et le Departement par nom
    exact (insensible à la casse) : si l'un des deux n'existe pas, la
    conversion échoue avec un message clair — c'est au RH de créer
    l'établissement/département correspondant avant de convertir.
    """
    demande = db.query(DemandeStage).filter(DemandeStage.id == demande_id).first()
    if demande is None:
        raise HTTPException(status_code=404, detail="Demande de stage non trouvée")

    if demande.stagiaire_id_cree is not None:
        raise HTTPException(
            status_code=400,
            detail="Cette demande a déjà été convertie en stagiaire",
        )

    etablissement = (
        db.query(Etablissement)
        .filter(Etablissement.nom.ilike(demande.etablissements.strip()))
        .first()
    )
    if etablissement is None:
        raise HTTPException(
            status_code=400,
            detail=(
                f"Aucun établissement nommé « {demande.etablissements} » n'existe. "
                "Créez-le d'abord (POST /etablissements) avant de convertir cette demande."
            ),
        )

    departement = (
        db.query(Departement)
        .filter(Departement.nom.ilike(demande.departements.strip()))
        .first()
    )
    if departement is None:
        raise HTTPException(
            status_code=400,
            detail=(
                f"Aucun département nommé « {demande.departements} » n'existe. "
                "Créez-le d'abord (POST /departements) avant de convertir cette demande."
            ),
        )

    try:
        date_debut_convertie = date.fromisoformat(demande.date_debut)
        date_fin_convertie = date.fromisoformat(demande.date_fin)
    except ValueError:
        raise HTTPException(
            status_code=400,
            detail="Les dates de la demande ne sont pas au format AAAA-MM-JJ, conversion impossible.",
        )

    nouveau_stagiaire = Stagiaire(
        prenom=demande.prenom,
        nom=demande.nom,
        email=demande.email,
        telephone=demande.telephone,
        cin=demande.cin,
        etablissement_id=etablissement.id,
        niveau_etudes=demande.niveau_etudes,
        specialisation=demande.specialisation,
        type_stage=demande.type_stage,
        departement_id=departement.id,
        encadrant_id=None,
        date_debut=date_debut_convertie,
        date_fin=date_fin_convertie,
        statut="en_cours",
        notifier_par_email=True,
    )
    db.add(nouveau_stagiaire)
    db.flush()  # attribue l'id avant de l'utiliser

    demande.statut = "acceptee"
    demande.stagiaire_id_cree = nouveau_stagiaire.id

    _log_activite(
        db,
        nouveau_stagiaire.id,
        f"Stagiaire créé à partir d'une candidature acceptée — {nouveau_stagiaire.prenom} {nouveau_stagiaire.nom}",
        nouveau_stagiaire.statut,
    )

    db.commit()
    db.refresh(nouveau_stagiaire)
    db.refresh(demande)
    return {"demande": demande, "stagiaire": nouveau_stagiaire}


# ==================== Routes API pour PFE Book ====================

@app.get("/sujets-pfe", response_model=list[SujetPFEResponse])
def liste_sujets_pfe(db: Session = Depends(get_db)):
    """Recupere tous les sujets PFE"""
    return db.query(SujetPFE).order_by(SujetPFE.annee.desc(), SujetPFE.reference).all()


@app.get("/sujets-pfe/ref/{reference}", response_model=SujetPFEResponse)
def get_sujet_by_ref(reference: str, db: Session = Depends(get_db)):
    """Recupere les details d'un sujet PFE par sa reference"""
    sujet = db.query(SujetPFE).filter(SujetPFE.reference == reference).first()
    if not sujet:
        raise HTTPException(status_code=404, detail="Sujet non trouve")
    return sujet


@app.post("/sujets-pfe", response_model=SujetPFEResponse)
def creer_sujet_pfe(sujet: SujetPFECreate, db: Session = Depends(get_db)):
    """Cree un nouveau sujet PFE (reserve aux RH/Admin)"""
    existing = db.query(SujetPFE).filter(SujetPFE.reference == sujet.reference).first()
    if existing:
        raise HTTPException(status_code=400, detail=f"La reference {sujet.reference} existe deja.")

    nouveau = SujetPFE(**sujet.model_dump())
    db.add(nouveau)
    db.commit()
    db.refresh(nouveau)
    return nouveau


@app.put("/sujets-pfe/{sujet_id}", response_model=SujetPFEResponse)
def modifier_sujet_pfe(sujet_id: int, sujet_maj: SujetPFEUpdate, db: Session = Depends(get_db)):
    """Modifie un sujet PFE existant"""
    sujet = db.query(SujetPFE).filter(SujetPFE.id == sujet_id).first()
    if not sujet:
        raise HTTPException(status_code=404, detail="Sujet non trouve")

    for champ, valeur in sujet_maj.model_dump(exclude_unset=True).items():
        setattr(sujet, champ, valeur)

    db.commit()
    db.refresh(sujet)
    return sujet


@app.delete("/sujets-pfe/{sujet_id}")
def supprimer_sujet_pfe(sujet_id: int, db: Session = Depends(get_db)):
    """Supprime un sujet PFE"""
    sujet = db.query(SujetPFE).filter(SujetPFE.id == sujet_id).first()
    if not sujet:
        raise HTTPException(status_code=404, detail="Sujet non trouve")

    db.delete(sujet)
    db.commit()
    return {"message": f"Sujet {sujet_id} supprime"}

@app.post("/envoyer-identifiants")
def envoyer_identifiants_api(request: EmailIdentifiantsRequest, db: Session = Depends(get_db)):
    """Envoie un email professionnel avec les identifiants de connexion via Resend."""
    
    if not RESEND_API_KEY:
        print("️ ATTENTION : RESEND_API_KEY non définie dans le fichier .env")
        raise HTTPException(status_code=503, detail="Clé API Resend non configurée")
    
    headers = {
        "Authorization": f"Bearer {RESEND_API_KEY}",
        "Content-Type": "application/json"
    }
    
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
    </head>
    <body style="margin: 0; padding: 0; font-family: 'Segoe UI', Arial, sans-serif; background-color: #F5F7FB;">
        <table width="100%" cellpadding="0" cellspacing="0" style="background-color: #F5F7FB; padding: 40px 0;">
            <tr>
                <td align="center">
                    <table width="600" cellpadding="0" cellspacing="0" style="background-color: #FFFFFF; border-radius: 16px; overflow: hidden; box-shadow: 0 4px 20px rgba(0,0,0,0.08);">
                        
                        <!-- En-tête bleu Hutchinson -->
                        <tr>
                            <td style="background: linear-gradient(135deg, #1D2B5B 0%, #26397A 100%); padding: 40px 30px; text-align: center;">
                                <h1 style="color: #FFFFFF; margin: 0; font-size: 28px; font-weight: 700; letter-spacing: 1px;">HUTCHINSON</h1>
                                <p style="color: rgba(255,255,255,0.8); margin: 10px 0 0 0; font-size: 14px;">Plateforme de gestion des stagiaires</p>
                            </td>
                        </tr>
                        
                        <!-- Corps de l'email -->
                        <tr>
                            <td style="padding: 40px 30px;">
                                <h2 style="color: #1D2B5B; margin-top: 0; font-size: 22px;">Bonjour {request.nom},</h2>
                                <p style="color: #4B5563; font-size: 15px; line-height: 1.6;">
                                    Votre compte stagiaire sur la plateforme de gestion <strong style="color: #1D2B5B;">Hutchinson</strong> a été créé avec succès par le service Ressources Humaines.
                                </p>
                                
                                <!-- Bloc identifiants -->
                                <div style="background-color: #F8FAFC; border-left: 4px solid #E31E24; border-radius: 8px; padding: 20px; margin: 25px 0;">
                                    <p style="margin: 0 0 12px 0; color: #1D2B5B; font-weight: 600; font-size: 14px;">🔐 VOS IDENTIFIANTS DE CONNEXION</p>
                                    <p style="margin: 8px 0; color: #4B5563; font-size: 14px;">
                                        <strong>Email :</strong><br>
                                        <span style="color: #1D2B5B; font-weight: 600;">{request.email}</span>
                                    </p>
                                    <p style="margin: 8px 0; color: #4B5563; font-size: 14px;">
                                        <strong>Mot de passe :</strong><br>
                                        <code style="background-color: #FFFFFF; padding: 6px 12px; border-radius: 4px; color: #E31E24; font-family: 'Courier New', monospace; font-weight: 700; font-size: 15px; letter-spacing: 1px;">{request.mot_de_passe}</code>
                                    </p>
                                </div>
                                
                                <!-- Bouton de connexion -->
                                <div style="text-align: center; margin: 30px 0;">
                                    <a href="http://localhost:5173/login" style="background-color: #E31E24; color: #FFFFFF; text-decoration: none; padding: 14px 32px; border-radius: 8px; font-weight: 600; font-size: 15px; display: inline-block;">
                                        Se connecter à la plateforme →
                                    </a>
                                </div>
                                
                                <!-- Avertissement sécurité -->
                                <div style="background-color: #FEF3C7; border-radius: 8px; padding: 15px; margin-top: 25px;">
                                    <p style="margin: 0; color: #92400E; font-size: 13px; line-height: 1.5;">
                                        ⚠️ <strong>Important :</strong> Pour des raisons de sécurité, nous vous recommandons de changer ce mot de passe lors de votre première connexion. Ne partagez jamais vos identifiants.
                                    </p>
                                </div>
                                
                                <p style="color: #6B7280; font-size: 14px; margin-top: 30px; line-height: 1.6;">
                                    Cordialement,<br>
                                    <strong style="color: #1D2B5B;">L'équipe Ressources Humaines</strong><br>
                                    Hutchinson
                                </p>
                            </td>
                        </tr>
                        
                        <!-- Pied de page -->
                        <tr>
                            <td style="background-color: #F8FAFC; padding: 20px 30px; text-align: center; border-top: 1px solid #E5E7EB;">
                                <p style="margin: 0; color: #9CA3AF; font-size: 12px;">
                                    Cet email a été envoyé automatiquement. Merci de ne pas y répondre.
                                </p>
                            </td>
                        </tr>
                    </table>
                </td>
            </tr>
        </table>
    </body>
    </html>
    """
    
    data = {
        "from": "Hutchinson RH <onboarding@resend.dev>",
        "to": [request.email],
        "subject": "🔐 Vos identifiants de connexion - Plateforme Hutchinson",
        "html": html_content
    }
    
    try:
        response = requests.post(
            "https://api.resend.com/emails",
            headers=headers,
            json=data,
            timeout=10
        )
        response.raise_for_status()
        
        print(f"✅ Email envoyé avec succès à {request.email}")
        return {"message": "Email envoyé avec succès", "email": request.email}
        
    except requests.exceptions.HTTPError as e:
        print(f"❌ Erreur HTTP Resend : {e.response.text}")
        raise HTTPException(status_code=500, detail=f"Erreur d'envoi : {e.response.text}")
    except Exception as e:
        print(f"❌ Erreur envoi email : {e}")
        return {"message": "Compte créé, mais échec de l'envoi de l'email", "email": request.email}